import streamlit as st
import pandas as pd
import numpy as np
import os
import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo


# ── 썸머 기준 ────────────────────────────────────────────────
MONTHLY_TARGET = 500_000
MONTHLY_HANWHA_MIN_PREMIUM = 50_000

READY_BONUS_RATES = [0, 15, 20, 25, 30]

SUMMER_GRADES = [
    ("HWARANG", 15_000_000),
    ("크라운", 10_000_000),
    ("트리플", 8_000_000),
    ("더블", 5_000_000),
    ("일반", 3_000_000),
]

TABLE_SEQ = 0


# ── 기본 유틸 ────────────────────────────────────────────────
def mark(ok: bool) -> str:
    return "✅" if ok else "❌"


def won(x) -> str:
    try:
        return f"{float(x):,.0f} 원"
    except Exception:
        return ""


def pct(x) -> str:
    try:
        return f"{float(x):,.0f} %"
    except Exception:
        return ""


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def standardize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    if "계약일" in df.columns and "계약일자" not in df.columns:
        df.rename(columns={"계약일": "계약일자"}, inplace=True)

    if "초회보험료" in df.columns and "보험료" not in df.columns:
        df.rename(columns={"초회보험료": "보험료"}, inplace=True)

    return df


def safe_table_name(base: str) -> str:
    name = re.sub(r"[^A-Za-z0-9_]", "_", str(base))

    if not re.match(r"^[A-Za-z_]", name):
        name = f"tbl_{name}"

    return name[:254]


def unique_sheet_name(wb, base, limit=31):
    name = str(base)[:limit] if base else "Sheet"

    if name not in wb.sheetnames:
        return name

    i = 2
    while True:
        suffix = f"_{i}"
        trunc = limit - len(suffix)
        cand = f"{name[:trunc]}{suffix}"

        if cand not in wb.sheetnames:
            return cand

        i += 1


def autosize_columns_full(ws, padding=5):
    for column_cells in ws.columns:
        max_len = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )

        ws.column_dimensions[column_cells[0].column_letter].width = max_len + padding


# ── 보험사 분류 ───────────────────────────────────────────────
def is_hanwha_life_series(ins: pd.Series) -> pd.Series:
    ins = ins.astype(str).str.strip()

    return (
        ins.str.contains("한화", na=False)
        & ins.str.contains("생명", na=False)
    )


def is_db_nonlife_series(ins: pd.Series) -> pd.Series:
    ins = ins.astype(str).str.strip()

    return (
        ins.str.contains("DB", case=False, na=False)
        & (
            ins.str.contains("손", na=False)
            | ins.str.contains("화재", na=False)
            | ins.str.contains("손해", na=False)
        )
    )


def is_kb_nonlife_series(ins: pd.Series) -> pd.Series:
    ins = ins.astype(str).str.strip()

    return (
        ins.str.contains("KB", case=False, na=False)
        & (
            ins.str.contains("손", na=False)
            | ins.str.contains("화재", na=False)
            | ins.str.contains("손해", na=False)
        )
    )


def is_hanwha_nonlife_series(ins: pd.Series) -> pd.Series:
    ins = ins.astype(str).str.strip()

    return (
        ins.str.contains("한화", na=False)
        & (
            ins.str.contains("손", na=False)
            | ins.str.contains("화재", na=False)
            | ins.str.contains("손해", na=False)
        )
        & ~is_hanwha_life_series(ins)
    )


def is_heungkuk_nonlife_series(ins: pd.Series) -> pd.Series:
    ins = ins.astype(str).str.strip()

    return (
        ins.str.contains("흥국", na=False)
        & (
            ins.str.contains("화재", na=False)
            | ins.str.contains("손", na=False)
            | ins.str.contains("손해", na=False)
        )
    )


def is_special_nonlife_series(ins: pd.Series) -> pd.Series:
    """
    썸머 우대 손해보험사:
    흥국화재, KB손해, 한화손해, DB손해
    """
    return (
        is_db_nonlife_series(ins)
        | is_kb_nonlife_series(ins)
        | is_hanwha_nonlife_series(ins)
        | is_heungkuk_nonlife_series(ins)
    )


def is_nonlife_series(ins: pd.Series) -> pd.Series:
    ins = ins.astype(str).str.strip()

    return (
        ins.str.contains("손해|손보|화재|해상", regex=True, na=False)
        | is_special_nonlife_series(ins)
    )


def is_life_series(ins: pd.Series) -> pd.Series:
    ins = ins.astype(str).str.strip()

    return (
        ins.str.contains("생명", na=False)
        | ins.str.contains("라이프", na=False)
    )


def is_other_life_series(ins: pd.Series) -> pd.Series:
    return is_life_series(ins) & ~is_hanwha_life_series(ins)


# ── 데이터 준비 ──────────────────────────────────────────────
def load_df(uploaded_file) -> pd.DataFrame:
    df = pd.read_excel(uploaded_file)
    df = normalize_columns(df)
    df = standardize_columns(df)
    return df


def exclude_contracts(df: pd.DataFrame):
    """
    제외 조건:
    - 일시납
    - 연금성 / 저축성
    - 철회 / 해약 / 실효
    """
    excluded_df = pd.DataFrame()

    needed = {"납입방법", "상품군2", "계약상태"}

    if needed.issubset(df.columns):
        tmp = df.copy()

        tmp["납입방법"] = tmp["납입방법"].astype(str).str.strip()
        tmp["상품군2"] = tmp["상품군2"].astype(str).str.strip()
        tmp["계약상태"] = tmp["계약상태"].astype(str).str.strip()

        is_lumpsum = tmp["납입방법"].str.contains("일시납", na=False)
        is_savings = tmp["상품군2"].str.contains("연금성|저축성", regex=True, na=False)
        is_cancelled = tmp["계약상태"].str.contains("철회|해약|실효", regex=True, na=False)

        is_excluded = is_lumpsum | is_savings | is_cancelled

        excluded_df = tmp[is_excluded].copy()
        df_valid = tmp[~is_excluded].copy()

        return df_valid, excluded_df

    return df.copy(), excluded_df


def build_excluded_with_reason(exdf: pd.DataFrame) -> pd.DataFrame:
    base_cols = [
        "수금자명",
        "계약일자",
        "보험사",
        "상품명",
        "납입기간",
        "보험료",
        "납입방법",
        "계약상태",
        "제외사유",
    ]

    if exdf is None or exdf.empty:
        return pd.DataFrame(columns=base_cols)

    tmp = standardize_columns(exdf.copy())

    def reason_row(row):
        reasons = []

        if "일시납" in str(row.get("납입방법", "")):
            reasons.append("일시납")

        product_group = str(row.get("상품군2", ""))
        if "연금성" in product_group or "저축성" in product_group:
            reasons.append("연금/저축성")

        status = str(row.get("계약상태", ""))
        if "철회" in status:
            reasons.append("철회")
        if "해약" in status:
            reasons.append("해약")
        if "실효" in status:
            reasons.append("실효")

        return " / ".join(reasons) if reasons else "제외 조건 미상"

    tmp["제외사유"] = tmp.apply(reason_row, axis=1)

    for col in base_cols:
        if col not in tmp.columns:
            tmp[col] = ""

    tmp["계약일자"] = pd.to_datetime(tmp["계약일자"], errors="coerce").dt.strftime("%Y-%m-%d")

    tmp["납입기간"] = pd.to_numeric(tmp["납입기간"], errors="coerce").apply(
        lambda x: f"{int(x)}년" if pd.notnull(x) else ""
    )

    tmp["보험료"] = pd.to_numeric(tmp["보험료"], errors="coerce").apply(
        lambda x: won(x) if pd.notnull(x) else ""
    )

    return tmp[base_cols]


def check_required_columns(df: pd.DataFrame):
    required_columns = {
        "수금자명",
        "계약일자",
        "보험사",
        "상품명",
        "납입기간",
        "보험료",
    }

    return required_columns - set(df.columns)


# ── 썸머 계산 ────────────────────────────────────────────────
def compute_summer(df: pd.DataFrame) -> pd.DataFrame:
    df = standardize_columns(df.copy())

    df["계약일자_raw"] = pd.to_datetime(df["계약일자"], errors="coerce")
    df["계약월"] = df["계약일자_raw"].dt.month

    df["보험료"] = pd.to_numeric(df["보험료"], errors="coerce").fillna(0)
    df["납입기간_num"] = pd.to_numeric(df["납입기간"], errors="coerce").fillna(0).astype(int)

    if "쉐어율" in df.columns:
        df["쉐어율"] = df["쉐어율"].apply(
            lambda x: float(str(x).replace("%", "").strip())
            if pd.notnull(x) and str(x).strip() != ""
            else np.nan
        )
    else:
        df["쉐어율"] = np.nan

    ins = df["보험사"].astype(str).str.strip()
    term = df["납입기간_num"]

    is_hanwha_life = is_hanwha_life_series(ins)
    is_special_nonlife = is_special_nonlife_series(ins)
    is_nonlife = is_nonlife_series(ins)
    is_other_nonlife = is_nonlife & ~is_special_nonlife
    is_other_life = is_other_life_series(ins)

    # ✅ 썸머 환산율
    # 손해보험
    # - 10년납 초과: 흥국/한화/KB/DB 250%, 이외 손해/화재 100%
    # - 10년납 이하: 흥국/한화/KB/DB 100%, 이외 손해/화재 50%
    #
    # 생명보험
    # - 10년납 초과: 한화생명 150%, 이외 생명보험 100%
    # - 10년납 이하: 한화생명 100%, 이외 생명보험 50%
    df["썸머율"] = np.select(
        [
            is_special_nonlife & (term > 10),
            is_special_nonlife & (term <= 10),
            is_other_nonlife & (term > 10),
            is_other_nonlife & (term <= 10),
            is_hanwha_life & (term > 10),
            is_hanwha_life & (term <= 10),
            is_other_life & (term > 10),
            is_other_life & (term <= 10),
        ],
        [
            250,
            100,
            100,
            50,
            150,
            100,
            100,
            50,
        ],
        default=0,
    ).astype(int)

    # 현재 기준: 보험료가 이미 쉐어율 반영된 값이라고 보고 그대로 사용
    df["실적보험료"] = df["보험료"]
    df["썸머환산금액"] = df["실적보험료"] * df["썸머율"] / 100

    return df


def check_monthly_requirements(dfin: pd.DataFrame):
    """
    월별 조건:
    1. 한화생명 5만원 이상 1건
    2. 환산업적 50만원 이상
    """
    if dfin.empty:
        return {
            "환산금액": 0,
            "한화생명5만": False,
            "환산50만": False,
            "월달성": False,
        }

    summer_sum = dfin["썸머환산금액"].sum()
    amount_ok = summer_sum >= MONTHLY_TARGET

    hanwha_ok = (
        is_hanwha_life_series(dfin["보험사"])
        & (pd.to_numeric(dfin["보험료"], errors="coerce").fillna(0) >= MONTHLY_HANWHA_MIN_PREMIUM)
    ).any()

    total_ok = amount_ok and hanwha_ok

    return {
        "환산금액": summer_sum,
        "한화생명5만": hanwha_ok,
        "환산50만": amount_ok,
        "월달성": total_ok,
    }


def get_summer_grade(total_amount: float):
    """
    7월 + 8월 합산 환산업적 기준 등급 산정.
    가장 높은 등급부터 체크.
    """
    for grade, target in SUMMER_GRADES:
        if total_amount >= target:
            return grade, target

    return "미달성", 0


def get_next_grade_gap(total_amount: float):
    ascending = [
        ("일반", 3_000_000),
        ("더블", 5_000_000),
        ("트리플", 8_000_000),
        ("크라운", 10_000_000),
        ("HWARANG", 15_000_000),
    ]

    for grade, target in ascending:
        if total_amount < target:
            return grade, target, target - total_amount

    return None, None, 0


def check_final_summer_requirements(
    july_df: pd.DataFrame,
    august_df: pd.DataFrame,
    ready_bonus_rate: float = 0,
):
    """
    1. 월별 필수조건은 보너스 전 금액 기준으로 판단
    2. 등급 판정은 레디포썸머 보너스 반영 후 금액 기준으로 판단
    """
    july_req = check_monthly_requirements(july_df)
    august_req = check_monthly_requirements(august_df)

    base_total_amount = july_req["환산금액"] + august_req["환산금액"]
    bonus_amount = base_total_amount * ready_bonus_rate / 100
    final_total_amount = base_total_amount + bonus_amount

    amount_grade, grade_target = get_summer_grade(final_total_amount)
    next_grade, next_target, next_gap = get_next_grade_gap(final_total_amount)

    monthly_all_ok = july_req["월달성"] and august_req["월달성"]

    # 금액 기준 등급과 최종 인정 등급을 분리
    if monthly_all_ok:
        final_grade = amount_grade
    else:
        final_grade = "필수조건 미충족"

    return {
        "7월": july_req,
        "8월": august_req,
        "기본합산환산금액": base_total_amount,
        "레디포썸머보너스율": ready_bonus_rate,
        "레디포썸머보너스금액": bonus_amount,
        "합산환산금액": final_total_amount,
        "월별필수조건": monthly_all_ok,
        "금액기준등급": amount_grade,
        "최종인정등급": final_grade,
        "달성기준금액": grade_target,
        "다음등급": next_grade,
        "다음등급기준": next_target,
        "다음등급부족금액": next_gap,
    }


# ── 화면 표시 ────────────────────────────────────────────────
def to_styled(dfin: pd.DataFrame) -> pd.DataFrame:
    df = dfin.copy()

    if df.empty:
        return pd.DataFrame(columns=[
            "계약월",
            "수금자명",
            "계약일자",
            "보험사",
            "상품명",
            "납입기간",
            "보험료",
            "쉐어율",
            "실적보험료",
            "썸머율",
            "썸머환산금액",
        ])

    df["계약일자"] = pd.to_datetime(df["계약일자"], errors="coerce").dt.strftime("%Y-%m-%d")

    df["납입기간"] = pd.to_numeric(df["납입기간"], errors="coerce").apply(
        lambda x: f"{int(x)}년" if pd.notnull(x) else ""
    )

    df["보험료"] = df["보험료"].map(won)
    df["쉐어율"] = df["쉐어율"].apply(lambda x: pct(x) if pd.notnull(x) else "")
    df["실적보험료"] = df["실적보험료"].map(won)
    df["썸머율"] = df["썸머율"].map(pct)
    df["썸머환산금액"] = df["썸머환산금액"].map(won)

    cols = [
        "계약월",
        "수금자명",
        "계약일자",
        "보험사",
        "상품명",
        "납입기간",
        "보험료",
        "쉐어율",
        "실적보험료",
        "썸머율",
        "썸머환산금액",
    ]

    return df[[c for c in cols if c in df.columns]]


def money_box(title, value, color="#ff9800"):
    return f"""
    <div style='border: 2px solid {color}; border-radius: 10px; padding: 18px; background-color: #fff8e1; margin-bottom: 12px;'>
        <h4 style='color:{color}; margin:0 0 8px 0;'>{title}</h4>
        <p style='font-size:20px; font-weight:bold; margin:0;'>{value:,.0f} 원</p>
    </div>
    """


def bonus_box(base_amount, bonus_rate, bonus_amount, final_amount):
    return f"""
    <div style='border: 2px solid #6f42c1; border-radius: 10px; padding: 18px; background-color: #f3ecff; margin-bottom: 12px;'>
        <h4 style='color:#6f42c1; margin:0 0 10px 0;'>🎁 레디포썸머 보너스 반영</h4>
        <p style='margin:4px 0;'><strong>기본 썸머 환산업적:</strong> {base_amount:,.0f} 원</p>
        <p style='margin:4px 0;'><strong>보너스율:</strong> {bonus_rate:.0f} %</p>
        <p style='margin:4px 0;'><strong>보너스 가산금액:</strong> {bonus_amount:,.0f} 원</p>
        <p style='font-size:20px; font-weight:bold; margin:10px 0 0 0; color:#6f42c1;'>
            보너스 반영 최종 환산업적: {final_amount:,.0f} 원
        </p>
    </div>
    """


def grade_box(final_grade, amount_grade, base_amount, bonus_rate, bonus_amount, final_amount, monthly_ok):
    if final_grade == "필수조건 미충족":
        color = "#b80000"
        bg = "#fdecea"
    elif final_grade == "미달성":
        color = "#b80000"
        bg = "#fdecea"
    elif final_grade in ["일반", "더블"]:
        color = "#0c6b2c"
        bg = "#e6f4ea"
    elif final_grade in ["트리플", "크라운"]:
        color = "#7a4b00"
        bg = "#fff4d6"
    else:
        color = "#4b0082"
        bg = "#f0e6ff"

    monthly_text = "월별 필수조건 충족" if monthly_ok else "월별 필수조건 미충족"

    return f"""
    <div style='border: 2px solid {color}; border-radius: 12px; padding: 20px; background-color: {bg}; margin-bottom: 16px;'>
        <h3 style='color:{color}; margin:0 0 10px 0;'>최종 인정 등급: {final_grade}</h3>
        <p style='margin:4px 0;'><strong>기본 썸머 환산업적:</strong> {base_amount:,.0f} 원</p>
        <p style='margin:4px 0;'><strong>레디포썸머 보너스율:</strong> {bonus_rate:.0f} %</p>
        <p style='margin:4px 0;'><strong>레디포썸머 보너스금액:</strong> {bonus_amount:,.0f} 원</p>
        <p style='font-size:20px; font-weight:bold; margin:8px 0;'>보너스 반영 최종 환산업적: {final_amount:,.0f} 원</p>
        <p style='font-weight:bold; margin:4px 0;'>금액 기준 등급: {amount_grade}</p>
        <p style='font-weight:bold; margin:4px 0;'>월별 필수조건: {monthly_text}</p>
    </div>
    """


def gap_box(title, amount):
    if amount > 0:
        color = "#e6f4ea"
        txt = "#0c6b2c"
        sym = f"+{amount:,.0f} 원 초과"
    elif amount < 0:
        color = "#fdecea"
        txt = "#b80000"
        sym = f"{amount:,.0f} 원 부족"
    else:
        color = "#f3f3f3"
        txt = "#000000"
        sym = "기준 달성"

    return f"""
    <div style='border: 1px solid {txt}; border-radius: 8px; background-color: {color}; padding: 12px; margin: 10px 0;'>
        <strong style='color:{txt};'>{title}: {sym}</strong>
    </div>
    """


def req_box(title, ok):
    color = "#e6f4ea" if ok else "#fdecea"
    txt = "#0c6b2c" if ok else "#b80000"
    mark_txt = "✅ 충족" if ok else "❌ 미충족"

    return f"""
    <div style='border: 1px solid {txt}; border-radius: 8px; background-color: {color}; padding: 12px; margin: 10px 0;'>
        <strong style='color:{txt};'>{title}: {mark_txt}</strong>
    </div>
    """


def make_collector_summary(july_df: pd.DataFrame, august_df: pd.DataFrame) -> pd.DataFrame:
    """
    수금자별 요약은 기본 환산업적 기준으로 표시.
    레디포썸머 보너스는 선택 수금자 화면에서 직접 선택 후 별도 반영.
    """
    all_df = pd.concat([july_df, august_df], ignore_index=True)

    rows = []

    if all_df.empty:
        return pd.DataFrame(columns=[
            "수금자명",
            "7월건수",
            "7월환산",
            "7월한화5만",
            "7월50만",
            "7월달성",
            "8월건수",
            "8월환산",
            "8월한화5만",
            "8월50만",
            "8월달성",
            "기본합산환산",
            "월별필수조건",
            "기본금액등급",
        ])

    for collector, sub in all_df.groupby("수금자명", dropna=False):
        july_sub = sub[sub["계약월"] == 7].copy()
        august_sub = sub[sub["계약월"] == 8].copy()

        result = check_final_summer_requirements(
            july_sub,
            august_sub,
            ready_bonus_rate=0,
        )

        rows.append({
            "수금자명": collector,
            "7월건수": len(july_sub),
            "7월환산": result["7월"]["환산금액"],
            "7월한화5만": mark(result["7월"]["한화생명5만"]),
            "7월50만": mark(result["7월"]["환산50만"]),
            "7월달성": mark(result["7월"]["월달성"]),
            "8월건수": len(august_sub),
            "8월환산": result["8월"]["환산금액"],
            "8월한화5만": mark(result["8월"]["한화생명5만"]),
            "8월50만": mark(result["8월"]["환산50만"]),
            "8월달성": mark(result["8월"]["월달성"]),
            "기본합산환산": result["기본합산환산금액"],
            "월별필수조건": mark(result["월별필수조건"]),
            "기본금액등급": result["금액기준등급"],
        })

    summary = pd.DataFrame(rows)
    return summary


def format_summary_for_display(summary: pd.DataFrame) -> pd.DataFrame:
    df = summary.copy()

    for col in ["7월환산", "8월환산", "기본합산환산"]:
        if col in df.columns:
            df[col] = df[col].map(won)

    return df


# ── 선택 수금자 필터 ─────────────────────────────────────────
def filter_by_collector(df: pd.DataFrame, selected_collector: str) -> pd.DataFrame:
    if selected_collector == "전체":
        return df.copy()

    return df[df["수금자명"].astype(str) == selected_collector].copy()


# ── 엑셀 출력 ────────────────────────────────────────────────
def write_table(ws, df_for_sheet: pd.DataFrame, start_row: int = 1, name_suffix: str = "A"):
    global TABLE_SEQ

    r_idx = start_row

    for r_idx, row in enumerate(dataframe_to_rows(df_for_sheet, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    end_col_letter = ws.cell(row=start_row, column=max(df_for_sheet.shape[1], 1)).column_letter
    last_row = r_idx if df_for_sheet.shape[0] > 0 else start_row

    TABLE_SEQ += 1
    display_name = safe_table_name(f"tbl_{ws.title}_{name_suffix}_{TABLE_SEQ}")

    table = Table(displayName=display_name, ref=f"A{start_row}:{end_col_letter}{last_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(table)

    autosize_columns_full(ws, padding=5)

    return last_row


def write_title(ws, row, title):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=13)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def write_final_result_block(ws, row, result):
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    fill = PatternFill("solid", fgColor="F2F2F2")

    rows = [
        ["7월 환산업적", won(result["7월"]["환산금액"])],
        ["7월 한화생명 5만원 이상 1건", mark(result["7월"]["한화생명5만"])],
        ["7월 환산업적 50만원 이상", mark(result["7월"]["환산50만"])],
        ["7월 조건 달성", mark(result["7월"]["월달성"])],
        ["8월 환산업적", won(result["8월"]["환산금액"])],
        ["8월 한화생명 5만원 이상 1건", mark(result["8월"]["한화생명5만"])],
        ["8월 환산업적 50만원 이상", mark(result["8월"]["환산50만"])],
        ["8월 조건 달성", mark(result["8월"]["월달성"])],
        ["기본 7월+8월 합산 환산업적", won(result["기본합산환산금액"])],
        ["레디포썸머 보너스율", f"{result['레디포썸머보너스율']:.0f} %"],
        ["레디포썸머 보너스금액", won(result["레디포썸머보너스금액"])],
        ["보너스 반영 최종 환산업적", won(result["합산환산금액"])],
        ["월별 필수조건", mark(result["월별필수조건"])],
        ["금액 기준 등급", result["금액기준등급"]],
        ["최종 인정 등급", result["최종인정등급"]],
    ]

    if result["다음등급"]:
        rows.append([
            f"다음 등급({result['다음등급']})까지 부족금액",
            won(result["다음등급부족금액"]),
        ])
    else:
        rows.append(["최고 등급 달성", "HWARANG"])

    for i, row_data in enumerate(rows, start=row):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

            if j == 1:
                cell.fill = fill
                cell.font = Font(bold=True)

    autosize_columns_full(ws, padding=5)

    return row + len(rows)


def build_workbook(
    df_all: pd.DataFrame,
    july_df: pd.DataFrame,
    august_df: pd.DataFrame,
    other_month_df: pd.DataFrame,
    summary: pd.DataFrame,
    result: dict,
    excluded_disp: pd.DataFrame,
):
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "요약"

    write_title(ws_summary, 1, "썸머 최종 결과")
    next_row = write_final_result_block(ws_summary, 2, result)

    write_title(ws_summary, next_row + 2, "수금자별 요약")
    next_row = write_table(
        ws_summary,
        format_summary_for_display(summary),
        start_row=next_row + 3,
        name_suffix="SUMMARY",
    )

    write_title(ws_summary, next_row + 2, "전체 상세")
    next_row = write_table(
        ws_summary,
        to_styled(df_all),
        start_row=next_row + 3,
        name_suffix="ALL_DETAIL",
    )

    ws_july = wb.create_sheet("7월")
    write_title(ws_july, 1, "7월 썸머 환산 결과")
    write_table(ws_july, to_styled(july_df), start_row=2, name_suffix="JULY_DETAIL")

    ws_august = wb.create_sheet("8월")
    write_title(ws_august, 1, "8월 썸머 환산 결과")
    write_table(ws_august, to_styled(august_df), start_row=2, name_suffix="AUGUST_DETAIL")

    if not other_month_df.empty:
        ws_other = wb.create_sheet("7월8월외")
        write_title(ws_other, 1, "7월/8월 외 계약")
        write_table(ws_other, to_styled(other_month_df), start_row=2, name_suffix="OTHER_MONTH")

    if excluded_disp is not None and not excluded_disp.empty:
        ws_ex = wb.create_sheet("제외계약")
        write_title(ws_ex, 1, "제외 계약")
        write_table(ws_ex, excluded_disp, start_row=2, name_suffix="EXCLUDED")

    return wb


# ── 탭 렌더링 함수 ───────────────────────────────────────────
def render_result_tabs(summary_df, july_df, august_df, other_month_df):
    tab1, tab2, tab3, tab4 = st.tabs(["🧮 수금자별 요약", "7월 상세", "8월 상세", "7월/8월 외"])

    with tab1:
        st.dataframe(format_summary_for_display(summary_df), use_container_width=True)

    with tab2:
        if july_df.empty:
            st.info("7월 계약이 없습니다.")
        else:
            st.dataframe(to_styled(july_df), use_container_width=True)

    with tab3:
        if august_df.empty:
            st.info("8월 계약이 없습니다.")
        else:
            st.dataframe(to_styled(august_df), use_container_width=True)

    with tab4:
        if other_month_df.empty:
            st.info("7월/8월 외 계약이 없습니다.")
        else:
            st.dataframe(to_styled(other_month_df), use_container_width=True)


# ── 메인 실행 ────────────────────────────────────────────────
def run():
    st.title("🌞 썸머 계산기")
    st.caption("엑셀 파일 하나를 업로드하면 계약일 기준으로 7월과 8월을 자동 분리하여 계산합니다.")

    with st.sidebar:
        st.subheader("🌞 월별 필수조건")
        st.markdown(
            f"""
            - 7월: 한화생명 **{MONTHLY_HANWHA_MIN_PREMIUM:,.0f}원 이상 1건**
            - 7월: 환산업적 **{MONTHLY_TARGET:,.0f}원 이상**
            - 8월: 한화생명 **{MONTHLY_HANWHA_MIN_PREMIUM:,.0f}원 이상 1건**
            - 8월: 환산업적 **{MONTHLY_TARGET:,.0f}원 이상**
            """
        )

        st.subheader("🎁 레디포썸머 보너스")
        st.markdown(
            """
            - 수금자 선택 후 보너스율 직접 선택
            - 선택 가능: 0%, 15%, 20%, 25%, 30%
            - 등급 판정은 보너스 반영 후 금액 기준
            - 월별 필수조건은 보너스 전 기준으로 판단
            """
        )

        st.subheader("🏆 최종 합산 등급")
        st.markdown(
            """
            - 일반: 300만원 이상
            - 더블: 500만원 이상
            - 트리플: 800만원 이상
            - 크라운: 1,000만원 이상
            - HWARANG: 1,500만원 이상
            """
        )

        st.subheader("📌 환산율")
        st.markdown(
            """
            **손해보험**
            - 10년납 초과: 흥국/한화/KB/DB 250%
            - 10년납 초과: 이외 손해/화재 100%
            - 10년납 이하: 흥국/한화/KB/DB 100%
            - 10년납 이하: 이외 손해/화재 50%

            **생명보험**
            - 10년납 초과: 한화생명 150%
            - 10년납 초과: 이외 생명보험 100%
            - 10년납 이하: 한화생명 100%
            - 10년납 이하: 이외 생명보험 50%
            """
        )

    uploaded_file = st.file_uploader(
        "📂 썸머 계산용 Excel 파일 업로드 (.xlsx)",
        type=["xlsx"],
        key="summer_one_file",
    )

    if uploaded_file is None:
        st.info("📤 7월과 8월 계약이 포함된 Excel 파일(.xlsx)을 업로드해주세요.")
        return

    base_filename = os.path.splitext(uploaded_file.name)[0]
    download_filename = f"{base_filename}_썸머환산결과.xlsx"

    try:
        raw = load_df(uploaded_file)
    except Exception as e:
        st.error(f"❌ 엑셀 파일을 읽는 중 오류가 발생했습니다: {e}")
        st.stop()

    df_valid, excluded_df = exclude_contracts(raw)
    excluded_disp = build_excluded_with_reason(excluded_df)

    missing = check_required_columns(df_valid)

    if missing:
        st.error(
            "❌ 업로드된 파일에 다음 항목이 필요합니다:\n\n"
            + ", ".join(sorted(missing))
        )
        st.stop()

    df = compute_summer(df_valid)

    invalid_dates = df[df["계약일자_raw"].isna()]
    if not invalid_dates.empty:
        st.warning(
            f"⚠️ {len(invalid_dates)}건의 계약일자가 날짜로 인식되지 않았습니다. "
            "계약일은 `2026-07-06`처럼 yyyy-mm-dd 형식으로 입력해주세요."
        )

    july_df = df[df["계약월"] == 7].copy()
    august_df = df[df["계약월"] == 8].copy()
    other_month_df = df[~df["계약월"].isin([7, 8])].copy()

    if july_df.empty:
        st.warning("⚠️ 계약일 기준 7월 계약이 없습니다.")

    if august_df.empty:
        st.warning("⚠️ 계약일 기준 8월 계약이 없습니다.")

    if not other_month_df.empty:
        st.info(
            f"ℹ️ 7월/8월 외 계약 {len(other_month_df)}건이 있습니다. "
            "이 계약들은 썸머 최종 조건 계산에서는 제외하고, 엑셀에는 별도 시트로 저장합니다."
        )

    # 전체 기준 결과: 보너스율 0% 기준
    total_result = check_final_summer_requirements(
        july_df,
        august_df,
        ready_bonus_rate=0,
    )
    total_summary = make_collector_summary(july_df, august_df)

    # 1. 제외 계약 보기 - 기본 펼침
    if excluded_disp is not None and not excluded_disp.empty:
        st.warning(
            f"⚠️ 제외된 계약 {len(excluded_disp)}건이 있습니다. "
            "제외 조건: 일시납 / 연금성·저축성 / 철회·해약·실효"
        )

        with st.expander("🚫 제외된 계약 보기", expanded=True):
            st.dataframe(excluded_disp, use_container_width=True)
    else:
        with st.expander("🚫 제외된 계약 보기", expanded=True):
            st.info("제외된 계약이 없습니다.")

    # 2. 전체 환산 결과
    st.subheader("📄 전체 환산 결과")
    render_result_tabs(
        summary_df=total_summary,
        july_df=july_df,
        august_df=august_df,
        other_month_df=other_month_df,
    )

    # 3. 수금자별 결과 확인
    st.subheader("👤 수금자별 결과 확인")

    collectors = ["전체"] + sorted(df["수금자명"].astype(str).dropna().unique().tolist())

    selected_collector = st.selectbox(
        "👤 확인할 수금자를 선택하세요.",
        collectors,
        index=0,
        key="summer_selected_collector",
    )

    ready_bonus_rate = st.selectbox(
        "🎁 레디포썸머 보너스율을 선택하세요.",
        READY_BONUS_RATES,
        index=0,
        format_func=lambda x: f"{x}%",
        key="summer_ready_bonus_rate",
    )

    selected_july_df = filter_by_collector(july_df, selected_collector)
    selected_august_df = filter_by_collector(august_df, selected_collector)
    selected_other_month_df = filter_by_collector(other_month_df, selected_collector)

    selected_summary = make_collector_summary(selected_july_df, selected_august_df)

    selected_result = check_final_summer_requirements(
        selected_july_df,
        selected_august_df,
        ready_bonus_rate=ready_bonus_rate,
    )

    st.markdown(f"### 📌 선택 기준: {selected_collector}")
    st.caption(f"레디포썸머 보너스율: {ready_bonus_rate}%")

    render_result_tabs(
        summary_df=selected_summary,
        july_df=selected_july_df,
        august_df=selected_august_df,
        other_month_df=selected_other_month_df,
    )

    # 4. 선택값 기준 월별 필수조건 체크
    st.subheader("✅ 월별 필수조건 체크")

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("### 7월")
        st.markdown(
            money_box("7월 환산업적", selected_result["7월"]["환산금액"]),
            unsafe_allow_html=True,
        )
        st.markdown(
            req_box(
                f"7월 한화생명 {MONTHLY_HANWHA_MIN_PREMIUM:,.0f}원 이상 1건",
                selected_result["7월"]["한화생명5만"],
            ),
            unsafe_allow_html=True,
        )
        st.markdown(
            req_box(
                f"7월 환산업적 {MONTHLY_TARGET:,.0f}원 이상",
                selected_result["7월"]["환산50만"],
            ),
            unsafe_allow_html=True,
        )
        st.markdown(
            req_box("7월 필수조건 전체", selected_result["7월"]["월달성"]),
            unsafe_allow_html=True,
        )

    with col2:
        st.markdown("### 8월")
        st.markdown(
            money_box("8월 환산업적", selected_result["8월"]["환산금액"]),
            unsafe_allow_html=True,
        )
        st.markdown(
            req_box(
                f"8월 한화생명 {MONTHLY_HANWHA_MIN_PREMIUM:,.0f}원 이상 1건",
                selected_result["8월"]["한화생명5만"],
            ),
            unsafe_allow_html=True,
        )
        st.markdown(
            req_box(
                f"8월 환산업적 {MONTHLY_TARGET:,.0f}원 이상",
                selected_result["8월"]["환산50만"],
            ),
            unsafe_allow_html=True,
        )
        st.markdown(
            req_box("8월 필수조건 전체", selected_result["8월"]["월달성"]),
            unsafe_allow_html=True,
        )

    st.markdown(
        req_box("7월·8월 월별 필수조건 전체", selected_result["월별필수조건"]),
        unsafe_allow_html=True,
    )

    # 5. 레디포썸머 보너스 반영 결과
    st.subheader("🎁 레디포썸머 보너스 반영 결과")

    st.markdown(
        bonus_box(
            selected_result["기본합산환산금액"],
            selected_result["레디포썸머보너스율"],
            selected_result["레디포썸머보너스금액"],
            selected_result["합산환산금액"],
        ),
        unsafe_allow_html=True,
    )

    # 6. 선택값 기준 썸머 최종 결과
    st.subheader("🏆 썸머 최종 결과")

    st.markdown(
        grade_box(
            selected_result["최종인정등급"],
            selected_result["금액기준등급"],
            selected_result["기본합산환산금액"],
            selected_result["레디포썸머보너스율"],
            selected_result["레디포썸머보너스금액"],
            selected_result["합산환산금액"],
            selected_result["월별필수조건"],
        ),
        unsafe_allow_html=True,
    )

    if selected_result["다음등급"]:
        st.markdown(
            gap_box(
                f"다음 등급 {selected_result['다음등급']}({selected_result['다음등급기준']:,.0f}원)까지",
                -selected_result["다음등급부족금액"],
            ),
            unsafe_allow_html=True,
        )
    else:
        st.success("🎉 최고 등급 HWARANG 기준을 달성했습니다.")

    # 7. 엑셀 다운로드
    wb = build_workbook(
        df_all=df,
        july_df=july_df,
        august_df=august_df,
        other_month_df=other_month_df,
        summary=total_summary,
        result=total_result,
        excluded_disp=excluded_disp,
    )

    excel_output = BytesIO()
    wb.save(excel_output)
    excel_output.seek(0)

    st.download_button(
        label="📥 썸머 환산 결과 엑셀 다운로드",
        data=excel_output,
        file_name=download_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )