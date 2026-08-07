from __future__ import annotations

# 전달용 파일: 보유계약 자동 연결·검토 흐름 적용본 v13

import hashlib
import io
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from difflib import SequenceMatcher
from typing import Any

import streamlit as st
import streamlit.components.v1 as components
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_PAYOUT_RATE = 65.0
FIRST_YEAR_HEADERS = {"1차년계", "1차년도합계", "1차년합계"}
TOTAL_HEADERS = {"총수수료", "총계", "총합계", "총수수료계"}
PRODUCT_HEADERS = {"상품명"}
PRODUCT_FALLBACK_HEADERS = {"구분"}


@dataclass(frozen=True)
class ProductRate:
    key: str
    source_type: str
    insurer: str
    product: str
    conditions: str
    first_year_rate: float
    total_rate: float
    sheet_name: str
    row_number: int

    @property
    def label(self) -> str:
        detail = f" · {self.conditions}" if self.conditions else ""
        return f"{self.product}{detail}"


@dataclass(frozen=True)
class HoldingContract:
    row_key: str
    source_type: str
    insurer_raw: str
    insurer: str
    policy_number: str
    product_raw: str
    customer: str
    premium: int
    payment_years: int | None
    payment_label: str
    contract_date: str
    contract_month: str
    status: str
    share_rate: float


def _normalize(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("計", "계")
    return re.sub(r"[\s\n\r\t:()\[\]·ㆍ_-]+", "", text).lower()


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _number(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if text in {"", "-"}:
            return None
        if text.endswith("%"):
            try:
                return float(text[:-1]) / 100
            except ValueError:
                return None
        try:
            return float(text)
        except ValueError:
            return None
    return None


def _effective_max_col(ws) -> int:
    """서식만 남아 XFD까지 확장된 시트의 불필요한 탐색을 막습니다."""
    upper = min(ws.max_column, 120)
    last = 1
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 40), max_col=upper):
        for cell in row:
            if cell.value not in (None, ""):
                last = max(last, cell.column)
    return min(max(last + 4, 20), upper)


def _is_header(value: Any, aliases: set[str]) -> bool:
    normalized = _normalize(value)
    return normalized in aliases


def _header_positions(ws, max_col: int) -> list[tuple[int, int, int, int]]:
    """상품명/1차년계/총수수료 열로 구성된 표 구간을 찾습니다."""
    positions: list[tuple[int, int, int, int]] = []
    for row_no in range(1, ws.max_row + 1):
        product_cols = [
            col for col in range(1, max_col + 1)
            if _is_header(ws.cell(row_no, col).value, PRODUCT_HEADERS)
        ]
        if not product_cols:
            product_cols = [
                col for col in range(1, max_col + 1)
                if _is_header(ws.cell(row_no, col).value, PRODUCT_FALLBACK_HEADERS)
            ]

        first_candidates: list[tuple[int, int]] = []
        total_candidates: list[tuple[int, int]] = []
        for header_row in range(row_no, min(row_no + 4, ws.max_row + 1)):
            for col in range(1, max_col + 1):
                value = ws.cell(header_row, col).value
                if _is_header(value, FIRST_YEAR_HEADERS):
                    first_candidates.append((header_row, col))
                if _is_header(value, TOTAL_HEADERS):
                    total_candidates.append((header_row, col))

        if not product_cols and first_candidates and total_candidates:
            product_cols = [1]
        if not product_cols:
            continue

        for product_col in product_cols:
            first_after = [item for item in first_candidates if item[1] > product_col]
            total_after = [item for item in total_candidates if item[1] > product_col]
            if not first_after or not total_after:
                continue
            first = min(first_after, key=lambda item: item[1])
            total = max(total_after, key=lambda item: item[1])
            if first[1] < total[1]:
                data_start = max(row_no, first[0], total[0]) + 1
                positions.append((data_start, product_col, first[1], total[1]))
                break
    return positions


def _condition_text(ws, row_no: int, product_col: int, first_col: int) -> str:
    ignored = {"-", "상품별상이", "해당없음"}
    values: list[str] = []
    for col in range(product_col + 1, first_col):
        value = ws.cell(row_no, col).value
        if isinstance(value, str) and value.startswith("="):
            continue
        text = _clean_text(value)
        if not text or text in ignored or text in values or _number(text) is not None:
            continue
        values.append(text)
    return " / ".join(values[:6])


def _source_payout_rate(formula_ws, value_ws, max_col: int) -> float | None:
    for row_no in range(1, min(formula_ws.max_row, 5) + 1):
        for col in range(1, max_col + 1):
            label = _normalize(formula_ws.cell(row_no, col).value)
            if label not in {"지급율", "지급률"}:
                continue
            for offset in range(1, 4):
                rate = _number(value_ws.cell(row_no, col + offset).value)
                if rate is not None:
                    return rate
    return None


def _extract_sheet(
    formula_ws,
    value_ws,
    source_type: str,
) -> tuple[list[ProductRate], list[str]]:
    insurer = formula_ws.title.strip()
    max_col = _effective_max_col(formula_ws)
    tables = _header_positions(formula_ws, max_col)
    results: list[ProductRate] = []
    warnings: list[str] = []
    source_payout = _source_payout_rate(formula_ws, value_ws, max_col)

    if not tables:
        return results, warnings
    if source_payout == 0:
        warnings.append(
            f"{insurer}: 원본 예시표의 지급율이 0%입니다. 해당 시트를 100%로 저장한 뒤 다시 올려 주세요."
        )
        return results, warnings

    for table_index, (data_start, product_col, first_col, total_col) in enumerate(tables):
        next_start = tables[table_index + 1][0] if table_index + 1 < len(tables) else formula_ws.max_row + 1
        end_row = next_start - 2
        current_product = ""

        for row_no in range(data_start, end_row + 1):
            raw_product = formula_ws.cell(row_no, product_col).value
            product_text = _clean_text(raw_product)

            if product_text:
                normalized_product = _normalize(product_text)
                if (
                    normalized_product in PRODUCT_HEADERS
                    or product_text.startswith("■")
                    or "수수료타입변경" in normalized_product
                ):
                    current_product = ""
                    continue
                current_product = product_text

            if not current_product:
                continue

            first_rate = _number(value_ws.cell(row_no, first_col).value)
            total_rate = _number(value_ws.cell(row_no, total_col).value)
            if first_rate is None or total_rate is None or first_rate < 0 or total_rate < 0:
                continue
            if first_rate == 0 and total_rate == 0:
                continue

            if source_payout not in (None, 0):
                first_rate /= source_payout
                total_rate /= source_payout

            conditions = _condition_text(formula_ws, row_no, product_col, first_col)
            identity = f"{source_type}|{insurer}|{current_product}|{conditions}|{row_no}"
            key = hashlib.sha1(identity.encode("utf-8")).hexdigest()[:16]
            results.append(
                ProductRate(
                    key=key,
                    source_type=source_type,
                    insurer=insurer,
                    product=current_product,
                    conditions=conditions,
                    first_year_rate=first_rate,
                    total_rate=total_rate,
                    sheet_name=formula_ws.title,
                    row_number=row_no,
                )
            )

    if tables and not results:
        warnings.append(f"{insurer}: 표는 찾았지만 계산된 수수료율을 읽지 못했습니다.")
    return results, warnings


@st.cache_data(show_spinner=False)
def parse_commission_workbook(file_bytes: bytes, source_type: str) -> tuple[list[dict], list[str]]:
    """예시표의 저장된 계산 결과를 읽습니다. 원본 파일은 변경하지 않습니다."""
    formula_book = load_workbook(io.BytesIO(file_bytes), data_only=False, read_only=False)
    value_book = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
    products: list[dict] = []
    warnings: list[str] = []

    for sheet_name in formula_book.sheetnames:
        if "변경" in sheet_name or sheet_name not in value_book.sheetnames:
            continue
        extracted, sheet_warnings = _extract_sheet(
            formula_book[sheet_name], value_book[sheet_name], source_type
        )
        products.extend(item.__dict__ for item in extracted)
        warnings.extend(sheet_warnings)

    formula_book.close()
    value_book.close()
    return products, warnings


def _to_product_rate(item: dict) -> ProductRate:
    return ProductRate(**item)


def _format_rate(multiplier: float) -> str:
    return f"{multiplier * 100:,.1f}%"


def _format_won(value: float) -> str:
    return f"{round(value):,}원"


def _compact_excel_product(contract: dict) -> str:
    """다운로드용으로 상품명과 세부 조건의 반복 문구를 정리합니다."""
    product_name, product_note = _product_display_parts(contract.get("product", ""))
    insurer = _clean_text(contract.get("insurer", ""))
    for token in ("(무배당)", "무배당", "(무)", insurer):
        if token:
            product_name = product_name.replace(token, "")
    product_name = re.sub(r"\((?:\s*|(?:20)?\d{2}[.\-]\d{2})\)", "", product_name)
    product_name = product_name.replace("_", " ")
    product_name = re.sub(r"\s+", " ", product_name).strip(" _·-/")

    raw_condition = " / ".join(
        part for part in (product_note, _clean_text(contract.get("conditions", ""))) if part
    )
    condition_parts: list[str] = []
    for part in re.split(r"\s*/\s*", raw_condition):
        cleaned = _clean_text(part)
        if not cleaned or cleaned in {"Y", "단일", "기본 조건"} or cleaned in condition_parts:
            continue
        condition_parts.append(cleaned)
    concise_condition = " / ".join(condition_parts[:4])
    return f"{product_name}\n{concise_condition}" if concise_condition else product_name


def _make_excel(
    contracts: list[dict], payout_rate: float, reference_month: str, excluded: list[dict]
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "수수료 계산"
    total_premium = sum(item["premium"] for item in contracts)
    total_first = sum(item["premium"] * item["first_year_rate"] * payout_rate for item in contracts)
    total_commission = sum(item["premium"] * item["total_rate"] * payout_rate for item in contracts)
    ws.append(["수수료 계산 결과"])
    ws.append(["수수료표 기준월", reference_month or "확인 필요", "공통 지급율", payout_rate])
    ws.append(["계약 수", len(contracts), "월보험료 합계", total_premium])
    ws.append(["예상 익월수당 합계", round(total_first), "예상 총수당 합계", round(total_commission)])
    ws.append([])
    headers = ["고객명", "증권번호", "보험회사", "상품 및 세부 조건", "월보험료", "모집 정보",
               "익월 수수료율", "총수수료율", "예상 익월수당", "예상 총수당"]
    ws.append(headers)

    for contract in contracts:
        first_rate = contract["first_year_rate"] * payout_rate
        total_rate = contract["total_rate"] * payout_rate
        premium = contract["premium"]
        product_detail = _compact_excel_product(contract)
        share_rate = contract.get("share_rate", 100.0)
        recruiter_type = contract.get("recruiter_type", "")
        recruiting = f"{share_rate:g}%"
        if share_rate < 100 and recruiter_type:
            recruiting += f" · {recruiter_type}"
        ws.append([
            contract.get("customer", ""),
            contract.get("policy_number", ""),
            contract["insurer"],
            product_detail,
            premium,
            recruiting,
            first_rate,
            total_rate,
            round(premium * first_rate),
            round(premium * total_rate),
        ])

    header_fill = PatternFill("solid", fgColor="2563D9")
    ws.merge_cells("A1:J1")
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1E3A8A")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for cell in ws[6]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws["D2"].number_format = "0%"
    for cell in (ws["D3"], ws["A4"], ws["C4"]):
        cell.font = Font(bold=True)
    ws["B3"].number_format = '0"건"'
    for cell in (ws["D3"], ws["B4"], ws["D4"]):
        cell.number_format = '#,##0"원"'
    for row in range(7, ws.max_row + 1):
        ws.cell(row, 5).number_format = '#,##0"원"'
        for col in range(7, 9):
            ws.cell(row, col).number_format = "0.0%"
        for col in range(9, 11):
            ws.cell(row, col).number_format = '#,##0"원"'
        ws.row_dimensions[row].height = 42

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = [15, 22, 18, 64, 18, 20, 18, 18, 21, 21]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A7"
    ws.auto_filter.ref = f"A6:J{ws.max_row}"
    ws.row_dimensions[1].height = 28

    review_ws = wb.create_sheet("검토 제외 계약")
    review_headers = ["고객명", "증권번호", "보험회사", "상품명", "계약상태", "제외 사유"]
    review_ws.append(review_headers)
    for item in excluded:
        review_ws.append([
            item.get("customer", ""), item.get("policy_number", ""), item.get("insurer", ""),
            _compact_excel_product({"product": item.get("product", ""), "insurer": item.get("insurer", "")}),
            item.get("status", ""), item.get("reason", ""),
        ])
    for cell in review_ws[1]:
        cell.fill = PatternFill("solid", fgColor="64748B")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for col, width in enumerate([15, 22, 18, 64, 15, 55], start=1):
        review_ws.column_dimensions[get_column_letter(col)].width = width
    for row in range(2, review_ws.max_row + 1):
        review_ws.row_dimensions[row].height = 36
    for row in review_ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    review_ws.freeze_panes = "A2"
    if review_ws.max_row > 1:
        review_ws.auto_filter.ref = review_ws.dimensions

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


INSURER_ALIASES = {
    "KB라이프생명": "KB라이프", "KB라이프": "KB라이프",
    "DB손해보험": "DB손보", "DB손보": "DB손보",
    "KB손해보험": "KB손보", "KB손보": "KB손보",
    "메리츠화재": "메리츠", "메리츠": "메리츠",
    "한화손해보험": "한화손보", "한화손보": "한화손보",
}


def _standard_insurer(value: Any) -> str:
    text = _clean_text(value)
    return INSURER_ALIASES.get(text, text)


def _source_type_from_insurer(insurer: str, insurer_code: str = "") -> str:
    if "생명" in insurer or "라이프" in insurer or insurer_code.upper().startswith("L"):
        return "생보"
    return "손보"


def _month_from_filename(filename: str) -> str:
    patterns = [r"(20\d{2})[._년\-\s]*(0?[1-9]|1[0-2])\s*월", r"(20\d{2})[._\-](0?[1-9]|1[0-2])"]
    for pattern in patterns:
        match = re.search(pattern, filename)
        if match:
            return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}"
    return ""


def _date_text(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = _clean_text(value)
    match = re.search(r"(20\d{2})[./\-년\s]*(\d{1,2})[./\-월\s]*(\d{1,2})?", text)
    if not match:
        return text
    year, month = int(match.group(1)), int(match.group(2))
    day = int(match.group(3) or 1)
    return f"{year:04d}-{month:02d}-{day:02d}"


def _holding_product_name(value: Any) -> str:
    text = _clean_text(value).lower()
    replacements = ("무배당", "(무)", "_무", "해약환급금", "해지환급금", "미지급형", "납입면제형")
    for token in replacements:
        text = text.replace(token, "")
    text = re.sub(r"\(?(?:20\d{2}|2\d)[.\-](?:0?[1-9]|1[0-2])\)?", "", text)
    text = re.sub(r"^kb|한화생명", "", text)
    return re.sub(r"[^0-9a-z가-힣]", "", text)


def _product_display_parts(product_name: str) -> tuple[str, str]:
    """상품명 셀 뒤에 붙은 납입 안내 문구를 화면용 상품명에서 분리합니다."""
    text = _clean_text(product_name)
    split_match = re.search(
        r"\s+(?=(?:\d+(?:\.\d+)?\s*구좌|\d+\s*년납(?:초과|이상|이하)))",
        text,
    )
    if not split_match:
        return text, ""
    return text[:split_match.start()].strip(), text[split_match.end():].strip()


def _condition_display(product: ProductRate) -> str:
    _, product_note = _product_display_parts(product.product)
    parts = [part for part in (product_note, product.conditions) if part]
    return " / ".join(dict.fromkeys(parts)) or "기본 조건"


@st.cache_data(show_spinner=False)
def parse_holding_workbook(file_bytes: bytes) -> list[dict]:
    """보유계약 장기 파일을 읽습니다. 잘못된 dimension=A1 파일도 처리합니다."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
    ws = wb[wb.sheetnames[0]]
    headers = {_normalize(cell.value): cell.column for cell in ws[1] if cell.value not in (None, "")}

    def value(row: int, *names: str) -> Any:
        for name in names:
            col = headers.get(_normalize(name))
            if col:
                return ws.cell(row, col).value
        return None

    results: list[dict] = []
    for row in range(2, ws.max_row + 1):
        policy_number = _clean_text(value(row, "증권번호"))
        product = _clean_text(value(row, "상품명"))
        insurer_raw = _clean_text(value(row, "보험사"))
        if not product and not policy_number:
            continue
        insurer = _standard_insurer(insurer_raw)
        date_value = _date_text(value(row, "계약일"))
        payment_year_number = _number(value(row, "납입기간"))
        payment_years = int(payment_year_number) if payment_year_number is not None else None
        payment_unit = _clean_text(value(row, "납입기간구분"))
        payment_label = f"{payment_years}{payment_unit}" if payment_years is not None else ""
        share_number = _number(value(row, "쉐어율"))
        share_rate = float(share_number if share_number is not None else 100.0)
        insurer_code = _clean_text(value(row, "보험사코드"))
        identity = f"{policy_number}|{product}|{date_value}|{row}"
        holding = HoldingContract(
            row_key=hashlib.sha1(identity.encode("utf-8")).hexdigest()[:16],
            source_type=_source_type_from_insurer(insurer, insurer_code),
            insurer_raw=insurer_raw,
            insurer=insurer,
            policy_number=policy_number,
            product_raw=product,
            customer=_clean_text(value(row, "계약자")),
            premium=int(_number(value(row, "계속보험료", "초회보험료")) or 0),
            payment_years=payment_years,
            payment_label=payment_label,
            contract_date=date_value,
            contract_month=date_value[:7] if re.match(r"20\d{2}-\d{2}", date_value) else "",
            status=_clean_text(value(row, "계약상태")) or "확인 필요",
            share_rate=share_rate,
        )
        results.append(holding.__dict__)
    wb.close()
    return results


def _payment_matches(product: ProductRate, years: int | None) -> bool:
    if years is None:
        return True
    condition = re.sub(r"\s+", "", product.conditions)
    if not condition:
        return True
    exact = re.search(rf"(?<!\d){years}년(?:납|갱신|만기)", condition)
    if exact:
        return True
    over = re.search(r"(\d+)년납(?:이상|↑)", condition)
    return bool(over and years >= int(over.group(1)))


def _rank_products(holding: dict, products: list[ProductRate]) -> list[tuple[float, ProductRate]]:
    source = _holding_product_name(holding["product_raw"])
    ranked: list[tuple[float, ProductRate]] = []
    for product in products:
        if product.source_type != holding["source_type"] or product.insurer != holding["insurer"]:
            continue
        target = _holding_product_name(product.product)
        if not source or not target:
            continue
        score = SequenceMatcher(None, source, target).ratio()
        if source in target or target in source:
            score = max(score, min(len(source), len(target)) / max(len(source), len(target)) + 0.08)
        if _payment_matches(product, holding.get("payment_years")):
            score += 0.04
        # 세만기·갱신형·간편형 등 핵심 유형이 서로 충돌하면 자동 확정을 방지합니다.
        raw = _normalize(holding["product_raw"])
        detail = _normalize(product.product + " " + product.conditions)
        for keyword in ("세만기", "연만기", "갱신형", "간편"):
            if keyword in raw and keyword not in detail:
                score -= 0.08
        ranked.append((score, product))
    ranked.sort(key=lambda item: item[0], reverse=True)
    return ranked


def _candidate_products(holding: dict, products: list[ProductRate]) -> list[ProductRate]:
    ranked = _rank_products(holding, products)
    if not ranked or ranked[0][0] < 0.58:
        return []
    best_name = _holding_product_name(ranked[0][1].product)
    same_product = [
        product for score, product in ranked
        if score >= max(0.55, ranked[0][0] - 0.14) and _holding_product_name(product.product) == best_name
    ]
    payment_filtered = [p for p in same_product if _payment_matches(p, holding.get("payment_years"))]
    candidates = payment_filtered or same_product
    unique: dict[tuple, ProductRate] = {}
    for product in candidates:
        key = (product.conditions, round(product.first_year_rate, 8), round(product.total_rate, 8))
        unique.setdefault(key, product)
    return list(unique.values())[:12]


def _review_candidate_products(holding: dict, products: list[ProductRate]) -> list[ProductRate]:
    """확인 화면에는 최상위 상품뿐 아니라 가까운 유사 상품·조건도 함께 보여줍니다."""
    primary = _candidate_products(holding, products)
    ranked = _rank_products(holding, products)
    if not ranked:
        return primary
    threshold = max(0.48, ranked[0][0] - 0.24)
    combined = primary + [product for score, product in ranked if score >= threshold][:30]
    unique: dict[tuple, ProductRate] = {}
    for product in combined:
        key = (
            _holding_product_name(product.product), product.conditions,
            round(product.first_year_rate, 8), round(product.total_rate, 8),
        )
        unique.setdefault(key, product)
    return list(unique.values())[:18]


def _auto_candidate(holding: dict, products: list[ProductRate]) -> ProductRate | None:
    ranked = _rank_products(holding, products)
    candidates = _candidate_products(holding, products)
    if not ranked or ranked[0][0] < 0.78 or not candidates:
        return None
    rate_pairs = {(round(p.first_year_rate, 8), round(p.total_rate, 8)) for p in candidates}
    if len(candidates) == 1 or len(rate_pairs) == 1:
        return candidates[0]
    return None


def _initialize_state() -> None:
    st.session_state.setdefault("commission_contracts", [])
    st.session_state.setdefault("commission_payout_rate", DEFAULT_PAYOUT_RATE)
    st.session_state.setdefault("commission_edit_index", None)
    st.session_state.setdefault("commission_edit_request", None)
    st.session_state.setdefault("commission_ratebook_signature", "")


def _reconnect_contract_rates(contracts: list[dict], products: list[ProductRate]) -> tuple[int, int]:
    """새 수수료표에서 기존 상품·세부 조건이 같은 계약의 요율을 다시 연결합니다."""
    updated_count = 0
    unresolved_count = 0
    for contract in contracts:
        same_product = [
            product for product in products
            if product.source_type == contract.get("source_type")
            and product.insurer == contract.get("insurer")
            and _holding_product_name(product.product) == _holding_product_name(contract.get("product", ""))
        ]
        same_condition = [
            product for product in same_product
            if _normalize(_condition_display(product)) == _normalize(contract.get("conditions", "") or "기본 조건")
        ]
        candidates = same_condition or same_product
        unique_rates = {
            (round(product.first_year_rate, 8), round(product.total_rate, 8)) for product in candidates
        }
        if not candidates or len(unique_rates) != 1:
            contract["rate_recheck_required"] = True
            unresolved_count += 1
            continue
        selected = candidates[0]
        contract.update({
            "product": selected.product, "conditions": selected.conditions,
            "first_year_rate": selected.first_year_rate, "total_rate": selected.total_rate,
            "sheet_name": selected.sheet_name, "row_number": selected.row_number,
            "rate_recheck_required": False,
        })
        updated_count += 1
    return updated_count, unresolved_count


def _contract_data(holding: dict, product: ProductRate, recruiter_type: str = "") -> dict:
    return {
        "customer": holding.get("customer", ""),
        "policy_number": holding.get("policy_number", ""),
        "insurer": product.insurer,
        "product": product.product,
        "conditions": product.conditions,
        "premium": int(holding.get("premium", 0)),
        "share_rate": float(holding.get("share_rate", 100.0)),
        "recruiter_type": recruiter_type,
        "contract_date": holding.get("contract_date", ""),
        "status": holding.get("status", ""),
        "source_type": product.source_type,
        "first_year_rate": product.first_year_rate,
        "total_rate": product.total_rate,
        "sheet_name": product.sheet_name,
        "row_number": product.row_number,
    }


def _holding_caption(holding: dict) -> str:
    policy = holding.get("policy_number") or "증권번호 없음"
    payment = f" · {holding['payment_label']}" if holding.get("payment_label") else ""
    return f"증권번호 {policy} · 월보험료 {_format_won(holding['premium'])}{payment}"


def _markdown_text(value: Any) -> str:
    """마스킹 이름의 ** 등이 Markdown 문법으로 해석되지 않도록 처리합니다."""
    return str(value or "").replace("\\", "\\\\").replace("*", "\\*").replace("_", "\\_")


def _render_manual_entry(all_products: list[ProductRate]) -> None:
    with st.expander("계약 직접 추가", expanded=False):
        if not all_products:
            st.info("생보 또는 손보 수수료 예시표를 먼저 올려 주세요.")
            return
        source_options = [source for source in ("생보", "손보") if any(p.source_type == source for p in all_products)]
        source_type = st.radio("보험 구분", source_options, horizontal=True, key="manual_source_type")
        insurers = sorted({p.insurer for p in all_products if p.source_type == source_type})
        insurer = st.selectbox("보험회사", insurers, index=None, key="manual_insurer")
        products = [p for p in all_products if p.source_type == source_type and p.insurer == insurer]
        product_groups: dict[str, list[ProductRate]] = defaultdict(list)
        for product in products:
            display_name, _ = _product_display_parts(product.product)
            product_groups[display_name].append(product)
        product_names = sorted(product_groups)
        product_name = st.selectbox("상품", product_names, index=None, key="manual_product")
        candidates = product_groups.get(product_name, [])
        if candidates:
            selected = st.selectbox(
                "납입기간 및 세부 조건", candidates, index=None,
                format_func=_condition_display, key="manual_condition",
                placeholder="납입기간과 세부 조건을 선택해 주세요.",
            )
        else:
            selected = None
            st.selectbox(
                "납입기간 및 세부 조건",
                ["먼저 상품을 선택해 주세요."],
                disabled=True,
                key="manual_condition_disabled",
            )
        col1, col2 = st.columns(2)
        customer = col1.text_input("고객명", key="manual_customer")
        policy = col2.text_input("증권번호", key="manual_policy")
        premium = st.number_input("월보험료", min_value=0, step=1000, value=0, format="%d", key="manual_premium")
        if st.button("직접 입력 계약 추가", type="primary", use_container_width=True):
            if selected is None or premium <= 0:
                st.warning("보험회사·상품·세부 조건과 월보험료를 확인해 주세요.")
            else:
                holding = {
                    "customer": customer.strip(), "policy_number": policy.strip(), "premium": int(premium),
                    "share_rate": 100.0, "contract_date": "", "status": "직접 등록",
                }
                st.session_state["commission_contracts"].append(_contract_data(holding, selected))
                st.rerun()


def _render_contract_editor(all_products: list[ProductRate]) -> None:
    edit_index = st.session_state.get("commission_edit_index")
    contracts = st.session_state["commission_contracts"]
    if not isinstance(edit_index, int) or not (0 <= edit_index < len(contracts)):
        return

    contract = contracts[edit_index]
    st.markdown(
        '<div id="commission-edit-anchor" style="scroll-margin-top:5rem;"></div>',
        unsafe_allow_html=True,
    )
    components.html(
        """
        <script>
        setTimeout(function () {
            const target = window.parent.document.getElementById('commission-edit-anchor');
            if (target) {
                target.scrollIntoView({ behavior: 'smooth', block: 'start' });
            }
        }, 180);
        </script>
        """,
        height=0,
    )
    st.info(f"{edit_index + 1}번 계약을 수정하고 있습니다.")
    with st.container(border=True):
        customer_col, policy_col = st.columns(2)
        customer = customer_col.text_input(
            "고객명", value=contract.get("customer", ""), key=f"edit_customer_{edit_index}"
        )
        policy_number = policy_col.text_input(
            "증권번호", value=contract.get("policy_number", ""), key=f"edit_policy_{edit_index}"
        )
        premium_col, share_col = st.columns(2)
        premium = premium_col.number_input(
            "월보험료", min_value=0, step=1000, value=int(contract.get("premium", 0)),
            format="%d", key=f"edit_premium_{edit_index}",
        )
        share_rate = share_col.number_input(
            "쉐어율 (%)", min_value=0.0, max_value=100.0,
            value=float(contract.get("share_rate", 100.0)), step=1.0, format="%.0f",
            key=f"edit_share_{edit_index}",
        )

        source_options = [source for source in ("생보", "손보") if any(p.source_type == source for p in all_products)]
        current_source = contract.get("source_type")
        source_index = source_options.index(current_source) if current_source in source_options else 0
        selected_source = st.radio(
            "보험 구분", source_options, index=source_index, horizontal=True,
            key=f"edit_source_{edit_index}",
        )
        insurer_options = sorted({p.insurer for p in all_products if p.source_type == selected_source})
        current_insurer = contract.get("insurer")
        insurer_index = insurer_options.index(current_insurer) if current_insurer in insurer_options else None
        selected_insurer = st.selectbox(
            "보험회사", insurer_options, index=insurer_index,
            placeholder="보험회사를 선택해 주세요.", key=f"edit_insurer_{edit_index}",
        )

        insurer_products = [
            product for product in all_products
            if product.source_type == selected_source and product.insurer == selected_insurer
        ]
        product_groups: dict[str, list[ProductRate]] = defaultdict(list)
        for product in insurer_products:
            display_name, _ = _product_display_parts(product.product)
            product_groups[display_name].append(product)
        product_names = sorted(product_groups)
        current_product_name, _ = _product_display_parts(contract.get("product", ""))
        product_index = product_names.index(current_product_name) if current_product_name in product_names else None
        selected_product_name = st.selectbox(
            "상품", product_names, index=product_index,
            placeholder="상품을 선택해 주세요.", key=f"edit_product_{edit_index}_{selected_insurer}",
        ) if product_names else None

        matching = product_groups.get(selected_product_name, [])
        current_position = next(
            (
                index for index, product in enumerate(matching)
                if product.sheet_name == contract.get("sheet_name") and product.row_number == contract.get("row_number")
            ),
            None,
        )
        if matching:
            selected_product = st.selectbox(
                "납입기간 및 세부 조건", matching,
                index=current_position,
                placeholder="납입기간과 세부 조건을 선택해 주세요.",
                format_func=_condition_display,
                key=f"edit_condition_{edit_index}_{hashlib.sha1(str(selected_product_name).encode()).hexdigest()[:8]}",
            )
        else:
            selected_product = None
            st.selectbox(
                "납입기간 및 세부 조건",
                ["먼저 상품을 선택해 주세요."],
                disabled=True,
                key=f"edit_condition_disabled_{edit_index}_{selected_insurer}",
            )

        recruiter_type = contract.get("recruiter_type", "")
        if share_rate < 100:
            recruiter_options = ["주모집", "공동모집"]
            recruiter_index = recruiter_options.index(recruiter_type) if recruiter_type in recruiter_options else None
            recruiter_type = st.selectbox(
                "모집 형태", recruiter_options, index=recruiter_index,
                placeholder="모집 형태를 선택해 주세요.", key=f"edit_recruiter_{edit_index}",
            ) or ""
        else:
            recruiter_type = ""

        save_col, cancel_col = st.columns([3, 1])
        if save_col.button("수정 완료", type="primary", use_container_width=True, key=f"save_edit_{edit_index}"):
            if premium <= 0:
                st.warning("월보험료를 확인해 주세요.")
            elif share_rate < 100 and not recruiter_type:
                st.warning("모집 형태를 선택해 주세요.")
            else:
                updated = dict(contract)
                updated.update({
                    "customer": customer.strip(), "policy_number": policy_number.strip(),
                    "premium": int(premium), "share_rate": float(share_rate),
                    "recruiter_type": recruiter_type,
                })
                if selected_product is not None:
                    updated.update({
                        "insurer": selected_product.insurer,
                        "product": selected_product.product,
                        "conditions": selected_product.conditions,
                        "source_type": selected_product.source_type,
                        "first_year_rate": selected_product.first_year_rate,
                        "total_rate": selected_product.total_rate,
                        "sheet_name": selected_product.sheet_name,
                        "row_number": selected_product.row_number,
                        "rate_recheck_required": False,
                    })
                contracts[edit_index] = updated
                st.session_state["commission_edit_index"] = None
                st.rerun()
        if cancel_col.button("취소", use_container_width=True, key=f"cancel_edit_{edit_index}"):
            st.session_state["commission_edit_index"] = None
            st.rerun()


def run() -> None:
    _initialize_state()

    st.title("수수료 계산기")
    st.caption("수수료 예시표와 보유계약 장기 파일을 연결해 계약별 예상 수당을 계산합니다.")

    with st.expander("① 수수료 예시표 불러오기", expanded=True):
        life_file = st.file_uploader(
            "생보 수수료 예시표", type=["xlsx"], key="commission_life_file"
        )
        nonlife_file = st.file_uploader(
            "손보 수수료 예시표", type=["xlsx"], key="commission_nonlife_file"
        )

    all_products: list[ProductRate] = []
    parse_warnings: list[str] = []
    reference_months: dict[str, str] = {}
    ratebook_hash = hashlib.sha256()
    for uploaded, source_type in ((life_file, "생보"), (nonlife_file, "손보")):
        if uploaded is None:
            continue
        try:
            uploaded_bytes = uploaded.getvalue()
            ratebook_hash.update(source_type.encode("utf-8"))
            ratebook_hash.update(uploaded_bytes)
            parsed, warnings = parse_commission_workbook(uploaded_bytes, source_type)
            all_products.extend(_to_product_rate(item) for item in parsed)
            parse_warnings.extend(warnings)
            reference_months[source_type] = _month_from_filename(uploaded.name)
        except Exception as exc:
            st.error(f"{source_type} 예시표를 읽지 못했습니다: {exc}")

    if all_products:
        insurer_count = len({product.insurer for product in all_products})
        month_text = " · ".join(
            f"{source} {month.replace('-', '년 ')}월" if month else f"{source} 기준월 확인 필요"
            for source, month in reference_months.items()
        )
        st.success(f"{month_text} · 보험회사 {insurer_count}개 · 수수료 조건 {len(all_products):,}개")
    else:
        st.info("생보 또는 손보 수수료 예시표를 올리면 상품을 선택할 수 있습니다.")

    for warning in parse_warnings:
        st.warning(warning)

    current_ratebook_signature = ratebook_hash.hexdigest() if all_products else ""
    saved_ratebook_signature = st.session_state.get("commission_ratebook_signature", "")
    contracts = st.session_state["commission_contracts"]
    if current_ratebook_signature and not saved_ratebook_signature:
        st.session_state["commission_ratebook_signature"] = current_ratebook_signature
    elif (
        current_ratebook_signature and saved_ratebook_signature
        and current_ratebook_signature != saved_ratebook_signature and contracts
    ):
        st.warning(
            "수수료 예시표가 변경되었습니다. 기존 수당 계산 대상 계약의 요율을 "
            "새 수수료표 기준으로 재검증해야 합니다."
        )
        reconnect_col, clear_col = st.columns(2)
        if reconnect_col.button("새 수수료표로 다시 연결", type="primary", use_container_width=True):
            updated, unresolved = _reconnect_contract_rates(contracts, all_products)
            st.session_state["commission_ratebook_signature"] = current_ratebook_signature
            st.session_state["commission_edit_index"] = None
            st.toast(f"{updated}건 재연결 · {unresolved}건 직접 확인 필요")
            st.rerun()
        if clear_col.button("기존 계약 초기화", use_container_width=True):
            st.session_state["commission_contracts"] = []
            st.session_state["commission_ratebook_signature"] = current_ratebook_signature
            st.session_state["commission_edit_index"] = None
            st.rerun()
        st.stop()
    elif current_ratebook_signature and current_ratebook_signature != saved_ratebook_signature:
        st.session_state["commission_ratebook_signature"] = current_ratebook_signature

    st.markdown("### ② 지급율 및 계약 불러오기")
    payout_rate_percent = st.number_input(
        "공통 지급율 (%)",
        min_value=0.0,
        max_value=100.0,
        value=float(round(st.session_state["commission_payout_rate"])),
        step=1.0,
        format="%.0f",
        help="변경한 지급율은 현재 수당 계산 대상 계약에 일괄 적용됩니다.",
    )
    st.session_state["commission_payout_rate"] = payout_rate_percent
    payout_rate = payout_rate_percent / 100
    holding_file = st.file_uploader(
        "보유계약관리 장기 엑셀", type=["xlsx"], key="commission_holding_file",
        help="계약상태가 정상이고 수수료표 기준월과 같은 계약을 우선 분석합니다.",
    )

    review_records: list[dict] = []
    if holding_file is not None and all_products:
        try:
            holdings = parse_holding_workbook(holding_file.getvalue())
        except Exception as exc:
            holdings = []
            st.error(f"보유계약 파일을 읽지 못했습니다: {exc}")

        registered_policies = {c.get("policy_number") for c in st.session_state["commission_contracts"] if c.get("policy_number")}
        automatic: list[tuple[dict, ProductRate]] = []
        needs_review: list[tuple[dict, list[ProductRate], str]] = []
        excluded: list[tuple[dict, str]] = []
        unmatched: list[tuple[dict, str]] = []
        already_registered = 0

        for holding in holdings:
            ref_month = reference_months.get(holding["source_type"], "")
            if holding.get("policy_number") and holding["policy_number"] in registered_policies:
                already_registered += 1
                continue
            if holding.get("status") != "정상":
                excluded.append((holding, f"계약상태가 {holding.get('status') or '확인 필요'}이므로 기본 제외"))
                continue
            if ref_month and holding.get("contract_month") and holding["contract_month"] != ref_month:
                excluded.append((holding, f"계약월 {holding['contract_month']} / 수수료표 기준월 {ref_month}"))
                continue
            candidates = _candidate_products(holding, all_products)
            if not candidates:
                unmatched.append((holding, "수수료표에서 일치하는 상품을 찾지 못함"))
                continue
            auto = _auto_candidate(holding, all_products)
            if auto is not None and holding.get("share_rate", 100.0) >= 100:
                automatic.append((holding, auto))
            else:
                reason_parts = []
                if auto is None:
                    reason_parts.append("세부 조건 확인")
                if holding.get("share_rate", 100.0) < 100:
                    reason_parts.append("모집 형태 확인")
                review_candidates = _review_candidate_products(holding, all_products)
                needs_review.append((holding, review_candidates, " · ".join(reason_parts)))

        st.markdown("### ③ 연결 결과 확인")
        metric_cols = st.columns(4)
        metric_cols[0].metric("전체", f"{len(holdings)}건")
        metric_cols[1].metric("자동 연결", f"{len(automatic)}건")
        metric_cols[2].metric("확인 필요", f"{len(needs_review)}건")
        metric_cols[3].metric("미연결·제외", f"{len(unmatched) + len(excluded)}건")
        if already_registered:
            st.caption(f"이미 등록된 증권번호 {already_registered}건은 중복 분석에서 제외했습니다.")

        pending: list[dict] = []
        with st.expander(f"자동 연결 완료 {len(automatic)}건", expanded=True):
            if not automatic:
                st.caption("자동 연결된 계약이 없습니다.")
            for holding, product in automatic:
                col1, col2 = st.columns([0.08, 0.92])
                selected = col1.checkbox("선택", value=True, key=f"auto_{holding['row_key']}", label_visibility="collapsed")
                with col2:
                    customer_name = _markdown_text(holding.get("customer") or "고객명 없음")
                    st.markdown(f"**{customer_name} · {product.insurer}**")
                    st.caption(_holding_caption(holding))
                    st.write(f"{product.product} · {product.conditions or '기본 조건'}")
                    reason = "상품명 일치"
                    if holding.get("payment_label"):
                        reason += f" · {holding['payment_label']} 조건 일치"
                    st.caption(f"자동 연결 근거: {reason}")
                if selected:
                    pending.append(_contract_data(holding, product))

        with st.expander(f"확인 필요 {len(needs_review)}건", expanded=bool(needs_review)):
            if not needs_review:
                st.caption("확인이 필요한 계약이 없습니다.")
            for holding, candidates, reason in needs_review:
                customer_name = _markdown_text(holding.get("customer") or "고객명 없음")
                st.markdown(f"**{customer_name} · {holding['insurer']}**")
                st.caption(f"{_holding_caption(holding)} · {reason}")
                st.write(f"보유계약 상품: {holding['product_raw']}")
                product_groups: dict[str, list[ProductRate]] = defaultdict(list)
                for candidate in candidates:
                    display_name, _ = _product_display_parts(candidate.product)
                    product_groups[display_name].append(candidate)
                all_product_names = list(product_groups)
                show_more = False
                if len(all_product_names) > 3:
                    show_more = st.checkbox(
                        f"다른 유사 상품 보기 ({len(all_product_names) - 3}개 더 있음)",
                        value=False, key=f"show_more_{holding['row_key']}",
                    )
                product_names = all_product_names if show_more else all_product_names[:3]
                selected_product_name = st.selectbox(
                    f"연결할 상품 · {'전체 후보' if show_more else '추천 후보'} {len(product_names)}개",
                    product_names,
                    index=0 if len(product_names) == 1 else None,
                    key=f"review_product_{holding['row_key']}_{'all' if show_more else 'recommended'}",
                    placeholder="수수료표의 상품을 선택해 주세요.",
                )
                selected_product = None
                if selected_product_name:
                    condition_candidates = product_groups[selected_product_name]
                    product_key = hashlib.sha1(selected_product_name.encode("utf-8")).hexdigest()[:8]
                    selected_product = st.selectbox(
                        f"납입기간 및 세부 조건 · 후보 {len(condition_candidates)}개",
                        condition_candidates,
                        index=0 if len(condition_candidates) == 1 else None,
                        key=f"review_condition_{holding['row_key']}_{product_key}",
                        placeholder="납입기간과 세부 조건을 선택해 주세요.",
                        format_func=lambda p: (
                            f"{_condition_display(p)} · "
                            f"익월 {_format_rate(p.first_year_rate * payout_rate)} / "
                            f"총 {_format_rate(p.total_rate * payout_rate)}"
                        ),
                    )
                else:
                    st.selectbox(
                        "납입기간 및 세부 조건",
                        ["먼저 상품을 선택해 주세요."],
                        disabled=True,
                        key=f"review_condition_disabled_{holding['row_key']}",
                    )
                recruiter_type = ""
                if holding.get("share_rate", 100.0) < 100:
                    recruiter_type = st.selectbox(
                        f"모집 형태 · 엑셀 쉐어율 {holding['share_rate']:g}%",
                        ["주모집", "공동모집"], index=None, key=f"recruiter_{holding['row_key']}",
                        placeholder="모집 형태를 선택해 주세요.",
                    ) or ""
                include = st.checkbox("이 계약 등록", value=True, key=f"review_include_{holding['row_key']}")
                ready = include and selected_product is not None and (
                    holding.get("share_rate", 100.0) >= 100 or recruiter_type
                )
                if ready:
                    pending.append(_contract_data(holding, selected_product, recruiter_type))
                    st.success("등록 준비 완료")
                elif not include:
                    review_records.append({**holding, "product": holding["product_raw"], "reason": "사용자가 등록 대상에서 제외"})
                else:
                    st.caption("상품·납입기간·모집 형태 중 필요한 항목을 선택해 주세요.")
                st.divider()

        if excluded:
            with st.expander(f"기준월·계약상태·중복으로 제외 {len(excluded)}건", expanded=False):
                st.info("기본적으로 제외됩니다. 필요한 경우에만 계약을 펼쳐 포함해 주세요.")
                for holding, reason in excluded:
                    customer_name = _markdown_text(holding.get("customer") or "고객명 없음")
                    st.markdown(f"**{customer_name} · {holding['insurer']}**")
                    st.caption(f"{_holding_caption(holding)} · {reason}")
                    include = st.checkbox("이번 계산에 포함", value=False, key=f"excluded_include_{holding['row_key']}")
                    if include:
                        candidates = _candidate_products(holding, all_products)
                        selected_product = st.selectbox(
                            "적용할 상품 및 조건", candidates, index=None, key=f"excluded_product_{holding['row_key']}",
                            format_func=lambda p: f"{p.product} · {p.conditions or '기본 조건'}",
                        ) if candidates else None
                        confirmed = st.checkbox(
                            "제외 사유를 확인했으며 이번 계산에 포함합니다.",
                            key=f"excluded_confirm_{holding['row_key']}",
                        )
                        if selected_product is not None and confirmed:
                            pending.append(_contract_data(holding, selected_product))
                        else:
                            review_records.append({**holding, "product": holding["product_raw"], "reason": reason})
                    else:
                        review_records.append({**holding, "product": holding["product_raw"], "reason": reason})
                    st.divider()

        if unmatched:
            with st.expander(f"연결되지 않은 계약 {len(unmatched)}건", expanded=bool(unmatched)):
                for holding, reason in unmatched:
                    customer_name = _markdown_text(holding.get("customer") or "고객명 없음")
                    st.markdown(f"**{customer_name} · {holding['insurer']}**")
                    st.caption(_holding_caption(holding))
                    st.write(f"{holding['product_raw']} · {reason}")
                    review_records.append({**holding, "product": holding["product_raw"], "reason": reason})

        if pending:
            if st.button(f"선택한 계약 {len(pending)}건 등록", type="primary", use_container_width=True):
                existing = {c.get("policy_number") for c in st.session_state["commission_contracts"] if c.get("policy_number")}
                added = 0
                for contract in pending:
                    if contract.get("policy_number") and contract["policy_number"] in existing:
                        continue
                    st.session_state["commission_contracts"].append(contract)
                    if contract.get("policy_number"):
                        existing.add(contract["policy_number"])
                    added += 1
                st.toast(f"계약 {added}건을 등록했습니다.")
                st.rerun()
        elif holdings:
            st.info("현재 등록할 수 있는 계약이 없습니다. 확인 필요 계약의 조건을 선택해 주세요.")

    _render_manual_entry(all_products)

    contracts = st.session_state["commission_contracts"]
    st.markdown("### ④ 수당 계산 대상 계약")
    if not contracts:
        st.info("추가된 계약이 없습니다.")
        return

    _render_contract_editor(all_products)

    calculation_contracts = [contract for contract in contracts if not contract.get("rate_recheck_required")]
    recheck_count = len(contracts) - len(calculation_contracts)
    if recheck_count:
        st.warning(
            f"새 수수료표에서 요율을 확정하지 못한 계약 {recheck_count}건은 합계와 다운로드에서 제외했습니다. "
            "해당 계약의 수정 버튼을 눌러 상품과 세부 조건을 다시 선택해 주세요."
        )

    total_premium = sum(contract["premium"] for contract in calculation_contracts)
    total_first = sum(
        contract["premium"] * contract["first_year_rate"] * payout_rate
        for contract in calculation_contracts
    )
    total_commission = sum(
        contract["premium"] * contract["total_rate"] * payout_rate
        for contract in calculation_contracts
    )
    metric_cols = st.columns(3)
    metric_cols[0].metric("월보험료 합계", _format_won(total_premium))
    metric_cols[1].metric("예상 익월수당", _format_won(total_first))
    metric_cols[2].metric("예상 총수당", _format_won(total_commission))

    header_columns = st.columns([3.6, 1, 1.15, 1.15, 1.25, 1.25, 1.05])
    for column, label in zip(
        header_columns, ("계약 정보", "월보험료", "익월 수수료율", "총 수수료율", "예상 익월수당", "예상 총수당", "관리"),
    ):
        column.caption(label)

    for index, contract in enumerate(contracts):
        first_rate = contract["first_year_rate"] * payout_rate
        total_rate = contract["total_rate"] * payout_rate
        expected_first = contract["premium"] * first_rate
        expected_total = contract["premium"] * total_rate
        product_detail = contract["product"]
        if contract["conditions"]:
            product_detail += f" · {contract['conditions']}"

        row_columns = st.columns([3.6, 1, 1.15, 1.15, 1.25, 1.25, 1.05])
        with row_columns[0]:
            customer_name = _markdown_text(contract.get("customer") or "고객명 없음")
            st.markdown(f"**{index + 1}. {customer_name}** · {contract['insurer']}")
            policy = contract.get("policy_number") or "증권번호 없음"
            recruiting = ""
            if contract.get("share_rate", 100) < 100:
                recruiting = f" · {contract['share_rate']:g}% · {contract.get('recruiter_type') or '모집 형태 확인'}"
            st.caption(f"증권번호 {policy}{recruiting} · {product_detail}")
            if contract.get("rate_recheck_required"):
                st.warning("요율 재확인 필요")
        row_columns[1].write(_format_won(contract["premium"]))
        row_columns[2].write(_format_rate(first_rate))
        row_columns[3].write(_format_rate(total_rate))
        row_columns[4].write(_format_won(expected_first))
        row_columns[5].write(_format_won(expected_total))
        with row_columns[6]:
            edit_col, delete_col = st.columns(2)
            if edit_col.button("수정", key=f"edit_commission_{index}", help="이 계약 수정"):
                st.session_state["commission_edit_index"] = index
                st.rerun()
            if delete_col.button("✕", key=f"delete_commission_{index}", help="이 계약 삭제"):
                contracts.pop(index)
                current_edit = st.session_state.get("commission_edit_index")
                if current_edit == index:
                    st.session_state["commission_edit_index"] = None
                elif isinstance(current_edit, int) and current_edit > index:
                    st.session_state["commission_edit_index"] = current_edit - 1
                st.rerun()

        if index < len(contracts) - 1:
            st.markdown(
                '<hr style="margin:.25rem 0 .45rem;border:0;border-top:1px solid rgba(128,128,128,.18);">',
                unsafe_allow_html=True,
            )

    st.divider()
    clear_col, download_col = st.columns([1, 2])
    with clear_col:
        if st.button("전체 계약 지우기", use_container_width=True):
            st.session_state["commission_contracts"] = []
            st.session_state["commission_edit_index"] = None
            st.rerun()
    with download_col:
        months = sorted({month for month in reference_months.values() if month})
        reference_month = ", ".join(months)
        excel_bytes = _make_excel(calculation_contracts, payout_rate, reference_month, review_records)
        st.download_button(
            "엑셀 다운로드",
            data=excel_bytes,
            file_name="수수료_계산결과.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )
