import hashlib
import re
from copy import copy
from datetime import datetime
from io import BytesIO
from pathlib import Path

import openpyxl
import streamlit as st
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .ui_components import page_header


DEFAULT_COVERAGES = [
    "질병사망",
    "재해(상해)사망",
    "질병후유장해3%일경우",
    "재해(상해)장해3%일경우",
    "일반암",
    "유사암",
    "표적항암약물허가치료비",
    "항암방사선약물치료비",
    "뇌혈관",
    "뇌졸중",
    "뇌출혈",
    "허혈성심장질환",
    "급성심근경색증",
    "질병수술",
    "질병종수술",
    "상해수술",
    "상해종수술",
    "뇌혈관질환수술",
    "허혈성심장질환수술",
    "질병입원",
    "상해입원",
    "간병인지원입원일-질병",
    "간병인지원입원일-상해",
    "상해간호간병통합입원일당",
    "질병간호간병통합입원일당",
    "교통사고처리지원금",
    "교통사고처리지원금(6주미만)",
    "변호사선임비용",
    "운전자벌금(대인)",
    "운전자벌금(대물)",
    "자동차사고부상위로금",
    "일상생활배상책임",
    "치아보철치료비",
    "치아보존치료비",
    "골절진단비",
    "질병입원(실손)",
    "질병통원(실손)",
    "상해입원(실손)",
    "상해통원(실손)",
]


DISPLAY_NAMES = {
    "질병사망": "질병 사망",
    "재해(상해)사망": "재해(상해) 사망",
    "질병후유장해3%일경우": "질병 후유장해 3%일 경우",
    "질병후유장해80%이상": "질병 후유장해 80% 이상",
    "재해(상해)장해3%일경우": "상해 후유장해 3%일 경우",
    "재해(상해)장해80%이상": "상해 후유장해 80% 이상",
    "고액암": "고액 암",
    "일반암": "일반 암",
    "이차암(재진단,계속암)": "이차암(재진단·계속암)",
    "유사암": "유사 암",
    "표적항암약물허가치료비": "표적항암약물허가치료비",
    "항암방사선약물치료비": "항암방사선·약물치료비",
    "뇌혈관": "뇌 혈관",
    "뇌졸중": "뇌 졸중",
    "뇌출혈": "뇌 출혈",
    "허혈성심장질환": "허혈성 심장 질환",
    "급성심근경색증": "급성 심근경색증",
    "중증치매": "중증 치매",
    "경증치매": "경증 치매",
    "장기간병요양진단(1급)": "장기요양 1등급",
    "장기간병요양진단(1,2급)": "장기요양 1~2등급",
    "장기간병요양진단(1,2,3급)": "장기요양 1~3등급",
    "장기간병요양진단(1,2,3,4급)": "장기요양 1~4등급",
    "암산정특례": "암 산정특례",
    "뇌혈관산정특례": "뇌혈관 산정특례",
    "심장질환산정특례": "심장질환 산정특례",
    "중증치매산정특례": "중증치매 산정특례",
    "질병수술": "질병 수술",
    "질병종수술": "질병 종 수술(1~5종)",
    "상해수술": "상해 수술",
    "상해종수술": "상해 종 수술(1~5종)",
    "암수술": "암 수술",
    "뇌혈관질환수술": "뇌혈관 질환 수술",
    "허혈성심장질환수술": "허혈성 심장 질환 수술",
    "질병입원": "질병 입원",
    "상해입원": "상해 입원",
    "간병인지원입원일-질병": "간병인 지원(질병)",
    "간병인지원입원일-상해": "간병인 지원(상해)",
    "암입원": "암 입원",
    "상해간호간병통합입원일당": "간호간병통합입원(상해)",
    "질병간호간병통합입원일당": "간호간병통합입원(질병)",
    "질병통원": "질병 통원",
    "암통원": "암 통원",
    "상해통원": "상해 통원",
    "치과통원": "치과 통원",
    "응급실내원비": "응급실 내원비",
    "교통사고처리지원금": "교통사고 처리 지원금",
    "교통사고처리지원금(6주미만)": "교통사고 처리 지원금(6주 미만)",
    "변호사선임비용": "변호사 선임 비용",
    "운전자벌금(대인)": "운전자 벌금(대인)",
    "운전자벌금(대물)": "운전자 벌금(대물)",
    "자동차사고부상위로금": "자동차사고 부상 위로금",
    "일상생활배상책임": "일상생활 배상책임",
    "치아보철치료비": "치아 보철 치료비",
    "치아보존치료비": "치아 보존 치료비",
    "화상진단비": "화상 진단비",
    "골절진단비": "골절 진단비",
    "깁스치료비": "깁스 치료비",
    "통풍진단비": "통풍 진단비",
    "대상포진진단비": "대상포진 진단비",
    "질병입원(실손)": "질병 입원(실손)",
    "질병통원(실손)": "질병 통원(실손)",
    "상해입원(실손)": "상해 입원(실손)",
    "상해통원(실손)": "상해 통원(실손)",
    "반려동물배상책임(대물)": "반려동물 배상책임(대물)",
    "반려동물배상책임(대인)": "반려동물 배상책임(대인)",
    "반려동물수술비(개)": "반려동물 수술비(개)",
    "반려동물입원비(개)": "반려동물 입원비(개)",
    "반려동물통원비(개)": "반려동물 통원비(개)",
}


GROUP_RULES = [
    ("사망", {"질병사망", "재해(상해)사망"}),
    (
        "후유\n장해",
        {
            "질병후유장해3%일경우",
            "질병후유장해80%이상",
            "재해(상해)장해3%일경우",
            "재해(상해)장해80%이상",
        },
    ),
    (
        "암\n보장",
        {
            "고액암",
            "일반암",
            "이차암(재진단,계속암)",
            "유사암",
            "표적항암약물허가치료비",
            "항암방사선약물치료비",
        },
    ),
    ("뇌\n보장", {"뇌혈관", "뇌졸중", "뇌출혈"}),
    ("심장\n보장", {"허혈성심장질환", "급성심근경색증"}),
    (
        "치매·\n장기요양",
        {
            "중증치매",
            "경증치매",
            "장기간병요양진단(1급)",
            "장기간병요양진단(1,2급)",
            "장기간병요양진단(1,2,3급)",
            "장기간병요양진단(1,2,3,4급)",
        },
    ),
    ("산정\n특례", {"암산정특례", "뇌혈관산정특례", "심장질환산정특례", "중증치매산정특례"}),
    (
        "수술",
        {
            "질병수술",
            "질병종수술",
            "상해수술",
            "상해종수술",
            "암수술",
            "뇌혈관질환수술",
            "허혈성심장질환수술",
        },
    ),
    ("입원", {"질병입원", "상해입원", "암입원"}),
    (
        "간호\n간병",
        {
            "간병인지원입원일-질병",
            "간병인지원입원일-상해",
            "상해간호간병통합입원일당",
            "질병간호간병통합입원일당",
        },
    ),
    ("통원·\n응급", {"질병통원", "암통원", "상해통원", "치과통원", "응급실내원비"}),
    (
        "운전자",
        {
            "교통사고처리지원금",
            "교통사고처리지원금(6주미만)",
            "변호사선임비용",
            "운전자벌금(대인)",
            "운전자벌금(대물)",
            "자동차사고부상위로금",
        },
    ),
    ("배상\n책임", {"일상생활배상책임"}),
    ("치아", {"치아보철치료비", "치아보존치료비"}),
    ("생활\n보장", {"화상진단비", "골절진단비", "깁스치료비", "통풍진단비", "대상포진진단비"}),
    ("실손", {"질병입원(실손)", "질병통원(실손)", "상해입원(실손)", "상해통원(실손)"}),
    (
        "반려\n동물",
        {
            "반려동물배상책임(대물)",
            "반려동물배상책임(대인)",
            "반려동물수술비(개)",
            "반려동물입원비(개)",
            "반려동물통원비(개)",
        },
    ),
]


COLORS = {
    "header": "DCE6F2",
    "premium": "FCD5B5",
    "cancer": "EBF1DE",
    "brain": "FDEADA",
    "heart": "E6E0EC",
    "white": "FFFFFF",
    "blue": "0000FF",
    "black": "000000",
    "line": "728197",
    "red": "FF0000",
}


THIN_SIDE = Side(style="thin", color=COLORS["line"])
MEDIUM_SIDE = Side(style="medium", color=COLORS["black"])
THICK_SIDE = Side(style="thick", color=COLORS["black"])
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


def _normalize_label(value: object) -> str:
    return re.sub(r"\s+", "", str(value or ""))


def _to_number(value: object) -> int | float:
    if isinstance(value, (int, float)):
        return value
    digits = re.sub(r"[^0-9.-]", "", str(value or ""))
    if not digits or digits in {"-", ".", "-."}:
        return 0
    number = float(digits)
    return int(number) if number.is_integer() else number


def _group_for(label: str) -> str:
    normalized = _normalize_label(label)
    for group_name, members in GROUP_RULES:
        if normalized in {_normalize_label(item) for item in members}:
            return group_name
    return "기타"


def _extract_customer_name(value: object) -> str:
    text = str(value or "고객").strip()
    name = re.split(r"[을를]\s*위한", text, maxsplit=1)[0].strip()
    return name or "고객"


def _extract_age(value: object) -> str:
    match = re.search(r"(\d+)\s*세", str(value or ""))
    return match.group(1) if match else ""


def parse_source_file(main_bytes: bytes) -> dict:
    workbook = openpyxl.load_workbook(BytesIO(main_bytes), data_only=True)
    required = ["계약사항", "상품별보장내용"]
    missing = [name for name in required if name not in workbook.sheetnames]
    if missing:
        raise ValueError("필수 시트 없음:" + ",".join(missing))

    contracts_ws = workbook["계약사항"]
    coverage_ws = workbook["상품별보장내용"]
    customer_name = _extract_customer_name(contracts_ws["B2"].value)
    age = _extract_age(contracts_ws["D2"].value)

    contract_columns = []
    for col in range(6, coverage_ws.max_column + 1):
        if coverage_ws.cell(2, col).value or coverage_ws.cell(3, col).value:
            contract_columns.append(col)

    if not contract_columns:
        raise ValueError("원본 파일에서 보험계약 정보를 찾을 수 없습니다.")

    contracts = []
    for index, col in enumerate(contract_columns):
        contract_row = 9 + index
        contracts.append(
            {
                "company": coverage_ws.cell(2, col).value or "",
                "product": coverage_ws.cell(3, col).value or "",
                "coverage_period": coverage_ws.cell(4, col).value or "",
                "payment_count": coverage_ws.cell(5, col).value or "",
                "payment_cycle": coverage_ws.cell(6, col).value or "",
                "monthly": _to_number(coverage_ws.cell(7, col).value),
                "total": _to_number(contracts_ws.cell(contract_row, 10).value),
                "paid": _to_number(contracts_ws.cell(contract_row, 11).value),
                "remaining": _to_number(contracts_ws.cell(contract_row, 12).value),
            }
        )

    coverages = []
    started = False
    for row in range(9, coverage_ws.max_row + 1):
        raw_label = coverage_ws.cell(row, 2).value
        if raw_label in (None, ""):
            if started:
                break
            continue
        started = True
        label = _normalize_label(raw_label)
        values = [_to_number(coverage_ws.cell(row, col).value) for col in contract_columns]
        coverages.append(
            {
                "label": label,
                "display": DISPLAY_NAMES.get(label, str(raw_label).strip()),
                "group": _group_for(label),
                "values": values,
            }
        )

    if not coverages:
        raise ValueError("원본 파일에서 보장항목을 찾을 수 없습니다.")

    return {
        "customer_name": customer_name,
        "age": age,
        "contracts": contracts,
        "coverages": coverages,
    }


def _set_outline(ws, min_row: int, max_row: int, min_col: int, max_col: int, side: Side) -> None:
    def replace_side(cell, **changes) -> None:
        border = cell.border
        cell.border = Border(
            left=changes.get("left", copy(border.left)),
            right=changes.get("right", copy(border.right)),
            top=changes.get("top", copy(border.top)),
            bottom=changes.get("bottom", copy(border.bottom)),
            diagonal=copy(border.diagonal),
            diagonal_direction=border.diagonal_direction,
            diagonalUp=border.diagonalUp,
            diagonalDown=border.diagonalDown,
            outline=border.outline,
            vertical=copy(border.vertical),
            horizontal=copy(border.horizontal),
        )

    for col in range(min_col, max_col + 1):
        top_cell = ws.cell(min_row, col)
        bottom_cell = ws.cell(max_row, col)
        replace_side(top_cell, top=side)
        replace_side(bottom_cell, bottom=side)
    for row in range(min_row, max_row + 1):
        left_cell = ws.cell(row, min_col)
        right_cell = ws.cell(row, max_col)
        replace_side(left_cell, left=side)
        replace_side(right_cell, right=side)


def _extract_logo(template_bytes: bytes) -> bytes:
    template_wb = openpyxl.load_workbook(BytesIO(template_bytes))
    template_ws = template_wb.active
    if not template_ws._images:
        raise ValueError("print.xlsx에서 로고 이미지를 찾을 수 없습니다.")
    return template_ws._images[0]._data()


def _configure_print(ws, contract_count: int, coverage_count: int, last_row: int, last_col: int) -> None:
    # 계약 열이 넓으면 가로, 계약이 적고 보장행이 많으면 세로가 더 크게 출력됩니다.
    orientation = "landscape" if contract_count >= 4 else "portrait"
    if contract_count == 3 and coverage_count <= 45:
        orientation = "landscape"

    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.orientation = orientation
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:{get_column_letter(last_col)}{last_row}"
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.page_margins.left = 0.15
    ws.page_margins.right = 0.15
    ws.page_margins.top = 0.2
    ws.page_margins.bottom = 0.2
    ws.page_margins.header = 0
    ws.page_margins.footer = 0
    ws.sheet_view.zoomScale = 70


def build_analysis_file(
    main_bytes: bytes,
    template_bytes: bytes,
    selected_labels: list[str] | None = None,
) -> tuple[bytes, str, str]:
    data = parse_source_file(main_bytes)
    available = {item["label"]: item for item in data["coverages"]}

    if selected_labels is None:
        selected_labels = DEFAULT_COVERAGES
    normalized_selection = {_normalize_label(label) for label in selected_labels}
    selected = [item for item in data["coverages"] if item["label"] in normalized_selection]
    if not selected:
        raise ValueError("출력할 보장항목을 한 개 이상 선택해 주세요.")

    workbook = Workbook()
    ws = workbook.active
    ws.title = "보장 분석"
    ws.sheet_view.showGridLines = False

    contract_count = len(data["contracts"])
    last_col = 3 + contract_count
    last_col_letter = get_column_letter(last_col)
    coverage_start = 11

    output_items = [
        {"label": "일반사망", "display": "일반 사망", "group": "사망", "values": [0] * contract_count},
        *selected,
        {"label": "기타", "display": "기타", "group": "기타", "values": [0] * contract_count},
    ]
    coverage_end = coverage_start + len(output_items) - 1

    normal_font = Font(name="나눔고딕", size=10, color=COLORS["black"])
    bold_font = Font(name="나눔고딕", size=10, bold=True, color=COLORS["black"])
    blue_font = Font(name="나눔고딕", size=11, bold=True, color=COLORS["blue"])
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("A1:C1")
    age_text = f" (보험연령:{data['age']}세)" if data["age"] else ""
    ws["A1"] = f"{data['customer_name']}님의 보장 분석{age_text}"
    ws["A1"].font = Font(name="나눔고딕", size=13, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="bottom")
    ws.row_dimensions[1].height = 82
    for col in range(1, last_col + 1):
        ws.cell(1, col).fill = PatternFill("solid", fgColor=COLORS["white"])
        ws.cell(1, col).border = THIN_BORDER
        ws.cell(1, col).alignment = center

    logo = XLImage(BytesIO(_extract_logo(template_bytes)))
    logo.width = 350
    logo.height = 43
    ws.add_image(logo, "A1")

    ws.merge_cells("A2:A3")
    ws.merge_cells("B2:B3")
    ws.merge_cells("C2:C3")
    ws["A2"] = "합 계"
    ws["B2"] = "구분"
    ws["C2"] = "보장명"
    ws["A2"].font = Font(name="나눔고딕", size=11, bold=True, color=COLORS["red"])

    for row in range(2, 4):
        for col in range(1, last_col + 1):
            cell = ws.cell(row, col)
            cell.fill = PatternFill("solid", fgColor=COLORS["header"])
            cell.border = THIN_BORDER
            cell.alignment = center
            if cell.coordinate != "A2":
                cell.font = bold_font

    for index, contract in enumerate(data["contracts"], start=4):
        ws.cell(2, index, contract["company"])
        ws.cell(3, index, contract["product"])
        ws.cell(2, index).font = Font(name="나눔고딕", size=10, bold=True, color="1F4E78")
        ws.cell(3, index).font = Font(name="나눔고딕", size=9, bold=True)
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 55

    meta_rows = [
        (4, "보장기간", "coverage_period"),
        (5, "납입횟수", "payment_count"),
        (6, "납입주기", "payment_cycle"),
        (7, "월보험료", "monthly"),
        (8, "납입완료", "paid"),
        (9, "납입예정", "remaining"),
        (10, "총보험료", "total"),
    ]
    for row, label, key in meta_rows:
        if row <= 6:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            ws.cell(row, 1, label)
        else:
            ws.cell(row, 1, sum(_to_number(contract[key]) for contract in data["contracts"]))
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            ws.cell(row, 2, label)

        fill_color = COLORS["header"] if row == 7 or row <= 6 else COLORS["premium"]
        for col in range(1, last_col + 1):
            cell = ws.cell(row, col)
            cell.fill = PatternFill("solid", fgColor=fill_color if col == 1 or col >= 4 else COLORS["header"])
            cell.border = THIN_BORDER
            cell.alignment = center
            cell.font = bold_font

        for index, contract in enumerate(data["contracts"], start=4):
            ws.cell(row, index, contract[key])
            if row >= 7:
                ws.cell(row, index).number_format = '#,##0"원"'
                ws.cell(row, index).font = blue_font
        if row >= 7:
            ws.cell(row, 1).number_format = '#,##0"원"'
            ws.cell(row, 1).font = blue_font

    group_ranges: list[tuple[int, int]] = []
    group_start = coverage_start
    current_group = output_items[0]["group"]
    for offset, item in enumerate(output_items):
        row = coverage_start + offset
        group = item["group"]
        if group != current_group:
            group_ranges.append((group_start, row - 1))
            group_start = row
            current_group = group

        section_color = COLORS["white"]
        if group == "암\n보장":
            section_color = COLORS["cancer"]
        elif group == "뇌\n보장":
            section_color = COLORS["brain"]
        elif group == "심장\n보장":
            section_color = COLORS["heart"]

        ws.cell(row, 1, sum(_to_number(value) for value in item["values"]))
        ws.cell(row, 2, group)
        ws.cell(row, 3, item["display"])
        for index, value in enumerate(item["values"], start=4):
            ws.cell(row, index, value)

        for col in range(1, last_col + 1):
            cell = ws.cell(row, col)
            cell.fill = PatternFill("solid", fgColor=COLORS["header"] if col == 2 else section_color)
            cell.border = THIN_BORDER
            cell.alignment = center
            cell.font = blue_font if col == 1 else bold_font
            if col == 1 or col >= 4:
                cell.number_format = '#,##0"만원";[Red]-#,##0"만원";;'
        ws.row_dimensions[row].height = 25
    group_ranges.append((group_start, coverage_end))

    for start, end in group_ranges:
        if end > start:
            ws.merge_cells(start_row=start, start_column=2, end_row=end, end_column=2)
        ws.cell(start, 2).alignment = center
        _set_outline(ws, start, end, 1, last_col, MEDIUM_SIDE)

    _set_outline(ws, 2, 3, 1, last_col, MEDIUM_SIDE)
    _set_outline(ws, 4, 6, 1, last_col, MEDIUM_SIDE)
    _set_outline(ws, 7, 10, 1, last_col, MEDIUM_SIDE)
    _set_outline(ws, 1, coverage_end, 1, last_col, THICK_SIDE)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 11
    ws.column_dimensions["C"].width = 31
    for col in range(4, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 19

    _configure_print(ws, contract_count, len(selected), coverage_end, last_col)
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    today = datetime.today().strftime("%Y%m%d")
    filename = f"{data['customer_name']}_보장분석_{today}.xlsx"
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue(), filename, data["customer_name"]


def make_input_signature(main_bytes: bytes, mode: str, selected_labels: list[str]) -> str:
    digest = hashlib.sha256()
    digest.update(main_bytes)
    digest.update(mode.encode("utf-8"))
    digest.update("|".join(selected_labels).encode("utf-8"))
    return digest.hexdigest()


def run() -> None:
    page_header(
        "고객 상담",
        "보장 분석 도우미",
        "전체 보장분석 원본을 고객 상담용 엑셀로 자동 정리합니다.",
        "▤",
    )

    with st.expander("사용 방법 안내"):
        st.markdown(
            """
            1. 전체 보장내용이 포함된 **컨설팅보장분석.xlsx** 파일을 업로드합니다.
            2. **간편모드**는 추천 기본 보장을 즉시 적용합니다.
            3. **개인모드**는 전체 보장 중 원하는 항목을 직접 선택합니다.
            4. **보장 분석 시작**을 누른 뒤 결과 엑셀을 다운로드합니다.

            - 결과물은 A3 용지에 한 페이지로 맞춰집니다.
            - 계약 수와 보장항목 수에 따라 가로·세로 방향을 자동으로 결정합니다.
            """
        )
        st.caption("버전 v2.0.0 · 제작 박병선 팀장")

    template_path = Path(__file__).resolve().with_name("print.xlsx")
    try:
        template_bytes = template_path.read_bytes()
        template_error = None
    except Exception as exc:
        template_bytes = b""
        template_error = str(exc)

    st.markdown("### 1. 전체 보장분석 원본")
    uploaded_main = st.file_uploader(
        "전체 보장내용이 포함된 컨설팅보장분석.xlsx 파일을 업로드하세요",
        type=["xlsx"],
        key="analyzer_v2_main_file",
    )

    parsed = None
    parse_error = None
    main_bytes = uploaded_main.getvalue() if uploaded_main else b""
    if uploaded_main:
        try:
            parsed = parse_source_file(main_bytes)
        except Exception as exc:
            parse_error = exc
            st.error(str(exc))

    st.markdown("### 2. 분석 모드")
    mode = st.radio(
        "분석 방식을 선택하세요",
        ["간편모드", "개인모드"],
        horizontal=True,
        key="analyzer_v2_mode",
    )

    selected_labels: list[str] = []
    if parsed:
        available_labels = [item["label"] for item in parsed["coverages"]]
        default_labels = [label for label in available_labels if label in set(DEFAULT_COVERAGES)]

        if mode == "간편모드":
            selected_labels = default_labels
            st.info(f"기본 보장 {len(selected_labels)}개가 자동으로 적용됩니다.")
            with st.expander("간편모드 적용 항목 보기"):
                st.write([DISPLAY_NAMES.get(label, label) for label in selected_labels])
        else:
            label_to_display = {label: DISPLAY_NAMES.get(label, label) for label in available_labels}
            selected_labels = st.multiselect(
                "출력할 보장항목을 선택하세요",
                options=available_labels,
                default=default_labels,
                format_func=lambda label: f"{_group_for(label).replace(chr(10), ' ')} · {label_to_display[label]}",
                key=f"analyzer_v2_selected_coverages_{hashlib.sha256(main_bytes).hexdigest()[:10]}",
            )
            st.caption(f"전체 {len(available_labels)}개 중 {len(selected_labels)}개 선택")
    elif not uploaded_main:
        st.caption("원본 파일을 업로드하면 선택 가능한 전체 보장항목이 표시됩니다.")

    if template_error:
        st.error("로고가 포함된 print.xlsx 파일을 찾을 수 없습니다. analyzer 파일과 같은 폴더에 두세요.")

    ready = bool(parsed) and bool(selected_labels) and bool(template_bytes) and parse_error is None
    signature = make_input_signature(main_bytes, mode, selected_labels) if ready else None

    st.markdown("### 3. 분석 실행")
    if st.button(
        "보장 분석 시작",
        type="primary",
        disabled=not ready,
        use_container_width=True,
        key="analyzer_v2_run",
    ):
        st.session_state.pop("analyzer_v2_result", None)
        st.session_state.pop("analyzer_v2_error", None)
        try:
            with st.spinner("고객 상담용 보장분석 엑셀을 만들고 있습니다..."):
                result_bytes, filename, customer_name = build_analysis_file(
                    main_bytes,
                    template_bytes,
                    selected_labels,
                )
            st.session_state["analyzer_v2_result"] = {
                "signature": signature,
                "bytes": result_bytes,
                "filename": filename,
                "customer_name": customer_name,
                "mode": mode,
                "coverage_count": len(selected_labels),
            }
        except Exception as exc:
            st.session_state["analyzer_v2_error"] = {
                "signature": signature,
                "message": str(exc),
                "detail": repr(exc),
            }

    error = st.session_state.get("analyzer_v2_error")
    if error and error.get("signature") == signature:
        st.error(error["message"])
        with st.expander("오류 상세 보기"):
            st.code(error["detail"])

    result = st.session_state.get("analyzer_v2_result")
    if result and result.get("signature") == signature:
        st.divider()
        st.success("보장 분석이 완료되었습니다.")
        col1, col2, col3 = st.columns(3)
        col1.metric("고객명", result["customer_name"])
        col2.metric("분석 모드", result["mode"])
        col3.metric("보장항목", f"{result['coverage_count']}개")
        st.download_button(
            "결과 엑셀 다운로드",
            data=result["bytes"],
            file_name=result["filename"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
            key="analyzer_v2_download",
        )


if __name__ == "__main__":
    run()
