import streamlit as st
import pandas as pd
import numpy as np
import os
import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo


# ── 컨벤션 기준 ──────────────────────────────────────────────
CONVENTION_GENERAL_TARGET = 1_800_000
CONVENTION_DOUBLE_TARGET = 3_600_000
CONVENTION_TRIPLE_TARGET = 5_400_000

# 기존 코드 호환용: 일반 달성 기준
CONVENTION_TARGET = CONVENTION_GENERAL_TARGET

CONVENTION_MIN_COUNT = 5
CONVENTION_HANWHA_MIN_PREMIUM = 50_000

CONVENTION_LEVELS = [
    ("트리플 달성", CONVENTION_TRIPLE_TARGET),
    ("더블 달성", CONVENTION_DOUBLE_TARGET),
    ("일반 달성", CONVENTION_GENERAL_TARGET),
]

TABLE_SEQ = 0


# ── 기본 유틸 ────────────────────────────────────────────────
def mark(ok: bool) -> str:
    return "✅" if ok else "❌"


def won(x) -> str:
    try:
        return f"{float(x):,.0f} 원"
    except Exception:
        return ""


def pct(x) -> str:
    try:
        return f"{float(x):,.0f} %"
    except Exception:
        return ""


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def standardize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    if "계약일" in df.columns and "계약일자" not in df.columns:
        df.rename(columns={"계약일": "계약일자"}, inplace=True)

    if "초회보험료" in df.columns and "보험료" not in df.columns:
        df.rename(columns={"초회보험료": "보험료"}, inplace=True)

    return df


def safe_table_name(base: str) -> str:
    name = re.sub(r"[^A-Za-z0-9_]", "_", str(base))

    if not re.match(r"^[A-Za-z_]", name):
        name = f"tbl_{name}"

    return name[:254]


def unique_sheet_name(wb, base, limit=31):
    name = str(base)[:limit] if base else "Sheet"

    if name not in wb.sheetnames:
        return name

    i = 2
    while True:
        suffix = f"_{i}"
        trunc = limit - len(suffix)
        cand = f"{name[:trunc]}{suffix}"

        if cand not in wb.sheetnames:
            return cand

        i += 1


def autosize_columns_full(ws, padding=5):
    for column_cells in ws.columns:
        max_len = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        ws.column_dimensions[column_cells[0].column_letter].width = max_len + padding


# ── 보험사 분류 ───────────────────────────────────────────────
def is_hanwha_life_series(ins: pd.Series) -> pd.Series:
    ins = ins.astype(str).str.strip()

    return (
        ins.str.contains("한화", na=False)
        & ins.str.contains("생명", na=False)
    )


def is_db_nonlife_series(ins: pd.Series) -> pd.Series:
    ins = ins.astype(str).str.strip()

    return (
        ins.str.contains("DB", case=False, na=False)
        & (
            ins.str.contains("손", na=False)
            | ins.str.contains("화재", na=False)
            | ins.str.contains("손해", na=False)
        )
    )


def is_kb_nonlife_series(ins: pd.Series) -> pd.Series:
    ins = ins.astype(str).str.strip()

    return (
        ins.str.contains("KB", case=False, na=False)
        & (
            ins.str.contains("손", na=False)
            | ins.str.contains("화재", na=False)
            | ins.str.contains("손해", na=False)
        )
    )


def is_hanwha_nonlife_series(ins: pd.Series) -> pd.Series:
    ins = ins.astype(str).str.strip()

    return (
        ins.str.contains("한화", na=False)
        & (
            ins.str.contains("손", na=False)
            | ins.str.contains("화재", na=False)
            | ins.str.contains("손해", na=False)
        )
        & ~is_hanwha_life_series(ins)
    )


def is_heungkuk_nonlife_series(ins: pd.Series) -> pd.Series:
    ins = ins.astype(str).str.strip()

    return (
        ins.str.contains("흥국", na=False)
        & (
            ins.str.contains("화재", na=False)
            | ins.str.contains("손", na=False)
            | ins.str.contains("손해", na=False)
        )
    )


def is_special_nonlife_series(ins: pd.Series) -> pd.Series:
    """
    컨벤션 우대 손해보험사:
    흥국화재, KB손해, 한화손해, DB손해
    """
    return (
        is_db_nonlife_series(ins)
        | is_kb_nonlife_series(ins)
        | is_hanwha_nonlife_series(ins)
        | is_heungkuk_nonlife_series(ins)
    )


def is_nonlife_series(ins: pd.Series) -> pd.Series:
    ins = ins.astype(str).str.strip()

    return (
        ins.str.contains("손해|손보|화재|해상", regex=True, na=False)
        | is_special_nonlife_series(ins)
    )


def is_life_series(ins: pd.Series) -> pd.Series:
    ins = ins.astype(str).str.strip()

    return (
        ins.str.contains("생명", na=False)
        | ins.str.contains("라이프", na=False)
    )


def is_other_life_series(ins: pd.Series) -> pd.Series:
    return is_life_series(ins) & ~is_hanwha_life_series(ins)


# ── 데이터 준비 ──────────────────────────────────────────────
def load_df(uploaded_file) -> pd.DataFrame:
    df = pd.read_excel(uploaded_file)
    df = normalize_columns(df)
    df = standardize_columns(df)
    return df


def exclude_contracts(df: pd.DataFrame):
    """
    제외 조건:
    - 일시납
    - 연금성 / 저축성
    - 철회 / 해약 / 실효
    """
    excluded_df = pd.DataFrame()

    needed = {"납입방법", "상품군2", "계약상태"}

    if needed.issubset(df.columns):
        tmp = df.copy()

        tmp["납입방법"] = tmp["납입방법"].astype(str).str.strip()
        tmp["상품군2"] = tmp["상품군2"].astype(str).str.strip()
        tmp["계약상태"] = tmp["계약상태"].astype(str).str.strip()

        is_lumpsum = tmp["납입방법"].str.contains("일시납", na=False)
        is_savings = tmp["상품군2"].str.contains("연금성|저축성", regex=True, na=False)
        is_cancelled = tmp["계약상태"].str.contains("철회|해약|실효", regex=True, na=False)

        is_excluded = is_lumpsum | is_savings | is_cancelled

        excluded_df = tmp[is_excluded].copy()
        df_valid = tmp[~is_excluded].copy()

        return df_valid, excluded_df

    return df.copy(), excluded_df


def build_excluded_with_reason(exdf: pd.DataFrame) -> pd.DataFrame:
    base_cols = [
        "수금자명",
        "계약일자",
        "보험사",
        "상품명",
        "납입기간",
        "보험료",
        "납입방법",
        "계약상태",
        "제외사유",
    ]

    if exdf is None or exdf.empty:
        return pd.DataFrame(columns=base_cols)

    tmp = standardize_columns(exdf.copy())

    def reason_row(row):
        reasons = []

        if "일시납" in str(row.get("납입방법", "")):
            reasons.append("일시납")

        product_group = str(row.get("상품군2", ""))
        if "연금성" in product_group or "저축성" in product_group:
            reasons.append("연금/저축성")

        status = str(row.get("계약상태", ""))
        if "철회" in status:
            reasons.append("철회")
        if "해약" in status:
            reasons.append("해약")
        if "실효" in status:
            reasons.append("실효")

        return " / ".join(reasons) if reasons else "제외 조건 미상"

    tmp["제외사유"] = tmp.apply(reason_row, axis=1)

    for col in base_cols:
        if col not in tmp.columns:
            tmp[col] = ""

    tmp["계약일자"] = pd.to_datetime(tmp["계약일자"], errors="coerce").dt.strftime("%Y-%m-%d")

    tmp["납입기간"] = pd.to_numeric(tmp["납입기간"], errors="coerce").apply(
        lambda x: f"{int(x)}년" if pd.notnull(x) else ""
    )

    tmp["보험료"] = pd.to_numeric(tmp["보험료"], errors="coerce").apply(
        lambda x: won(x) if pd.notnull(x) else ""
    )

    return tmp[base_cols]


def check_required_columns(df: pd.DataFrame):
    required_columns = {
        "수금자명",
        "계약일자",
        "보험사",
        "상품명",
        "납입기간",
        "보험료",
    }

    return required_columns - set(df.columns)


# ── 컨벤션 계산 ──────────────────────────────────────────────
def compute_convention(df: pd.DataFrame) -> pd.DataFrame:
    df = standardize_columns(df.copy())

    df["계약일자_raw"] = pd.to_datetime(df["계약일자"], errors="coerce")
    df["보험료"] = pd.to_numeric(df["보험료"], errors="coerce").fillna(0)
    df["납입기간_num"] = pd.to_numeric(df["납입기간"], errors="coerce").fillna(0).astype(int)

    if "쉐어율" in df.columns:
        df["쉐어율"] = df["쉐어율"].apply(
            lambda x: float(str(x).replace("%", "").strip())
            if pd.notnull(x) and str(x).strip() != ""
            else np.nan
        )
    else:
        df["쉐어율"] = np.nan

    ins = df["보험사"].astype(str).str.strip()
    term = df["납입기간_num"]

    is_hanwha_life = is_hanwha_life_series(ins)
    is_special_nonlife = is_special_nonlife_series(ins)
    is_nonlife = is_nonlife_series(ins)
    is_other_nonlife = is_nonlife & ~is_special_nonlife
    is_other_life = is_other_life_series(ins)

    # ✅ 컨벤션 환산율
    # 1. 한화생명: 납입기간 상관없이 150%
    # 2. 흥국화재, KB손해, 한화손해, DB손해: 300%
    # 3. 그 외 손해보험: 200%
    # 4. 그 외 생명보험 10년납 미만: 50%
    # 5. 그 외 생명보험 10년납 이상: 100%
    df["컨벤션율"] = np.select(
        [
            is_hanwha_life,
            is_special_nonlife,
            is_other_nonlife,
            is_other_life & (term < 10),
            is_other_life & (term >= 10),
        ],
        [
            150,
            300,
            200,
            50,
            100,
        ],
        default=0,
    ).astype(int)

    # 현재 기준: 보험료가 이미 쉐어율 반영된 값이라고 보고 그대로 사용
    df["실적보험료"] = df["보험료"]
    df["컨벤션환산금액"] = df["실적보험료"] * df["컨벤션율"] / 100

    return df


def get_amount_level(conv_sum: float) -> str:
    """
    금액만 기준으로 봤을 때의 수준.
    실제 달성등급과 다를 수 있음.
    """
    for level_name, target in CONVENTION_LEVELS:
        if conv_sum >= target:
            return level_name

    return "미달성"


def get_final_level(conv_sum: float, count_ok: bool, hanwha_ok: bool) -> str:
    """
    최종 달성등급.
    계약 5건 이상 + 한화생명 5만 원 이상 1건이 필수 조건.
    필수 조건 미충족 시 금액이 높아도 달성 인정하지 않음.
    """
    required_ok = count_ok and hanwha_ok

    if not required_ok:
        if conv_sum >= CONVENTION_GENERAL_TARGET:
            return "필수조건 미충족"
        return "미달성"

    for level_name, target in CONVENTION_LEVELS:
        if conv_sum >= target:
            return level_name

    return "미달성"


def check_convention_requirements(dfin: pd.DataFrame):
    if dfin.empty:
        return {
            "컨벤션환산금액": 0,
            "금액수준": "미달성",
            "달성등급": "미달성",
            "환산180만": False,
            "일반달성": False,
            "더블달성": False,
            "트리플달성": False,
            "건수5건": False,
            "한화생명5만": False,
            "필수조건": False,
            "전체달성": False,
        }

    conv_sum = dfin["컨벤션환산금액"].sum()

    general_amount_ok = conv_sum >= CONVENTION_GENERAL_TARGET
    double_amount_ok = conv_sum >= CONVENTION_DOUBLE_TARGET
    triple_amount_ok = conv_sum >= CONVENTION_TRIPLE_TARGET

    count_ok = len(dfin) >= CONVENTION_MIN_COUNT

    hanwha_ok = (
        is_hanwha_life_series(dfin["보험사"])
        & (
            pd.to_numeric(dfin["보험료"], errors="coerce").fillna(0)
            >= CONVENTION_HANWHA_MIN_PREMIUM
        )
    ).any()

    required_ok = count_ok and hanwha_ok

    final_general_ok = required_ok and general_amount_ok
    final_double_ok = required_ok and double_amount_ok
    final_triple_ok = required_ok and triple_amount_ok

    amount_level = get_amount_level(conv_sum)
    final_level = get_final_level(conv_sum, count_ok, hanwha_ok)

    return {
        "컨벤션환산금액": conv_sum,
        "금액수준": amount_level,
        "달성등급": final_level,
        "환산180만": final_general_ok,
        "일반달성": final_general_ok,
        "더블달성": final_double_ok,
        "트리플달성": final_triple_ok,
        "건수5건": count_ok,
        "한화생명5만": hanwha_ok,
        "필수조건": required_ok,
        "전체달성": final_general_ok,
    }


# ── 화면 표시 ────────────────────────────────────────────────
def to_styled(dfin: pd.DataFrame) -> pd.DataFrame:
    df = dfin.copy()

    df["계약일자"] = pd.to_datetime(df["계약일자"], errors="coerce").dt.strftime("%Y-%m-%d")

    df["납입기간"] = pd.to_numeric(df["납입기간"], errors="coerce").apply(
        lambda x: f"{int(x)}년" if pd.notnull(x) else ""
    )

    df["보험료"] = df["보험료"].map(won)
    df["쉐어율"] = df["쉐어율"].apply(lambda x: pct(x) if pd.notnull(x) else "")
    df["실적보험료"] = df["실적보험료"].map(won)
    df["컨벤션율"] = df["컨벤션율"].map(pct)
    df["컨벤션환산금액"] = df["컨벤션환산금액"].map(won)

    cols = [
        "수금자명",
        "계약일자",
        "보험사",
        "상품명",
        "납입기간",
        "보험료",
        "쉐어율",
        "실적보험료",
        "컨벤션율",
        "컨벤션환산금액",
    ]

    return df[[c for c in cols if c in df.columns]]


def money_box(title, value, color="#1f77b4"):
    return f"""
    <div style='border: 2px solid {color}; border-radius: 10px; padding: 18px; background-color: #f7faff; margin-bottom: 12px;'>
        <h4 style='color:{color}; margin:0 0 8px 0;'>{title}</h4>
        <p style='font-size:20px; font-weight:bold; margin:0;'>{value:,.0f} 원</p>
    </div>
    """


def level_box(level_name):
    if level_name == "트리플 달성":
        color = "#7b2cbf"
        bg = "#f3e8ff"
        icon = "🏆"
    elif level_name == "더블 달성":
        color = "#0b5394"
        bg = "#e8f1ff"
        icon = "🥈"
    elif level_name == "일반 달성":
        color = "#0c6b2c"
        bg = "#e6f4ea"
        icon = "✅"
    elif level_name == "필수조건 미충족":
        color = "#b85c00"
        bg = "#fff4e5"
        icon = "⚠️"
    else:
        color = "#b80000"
        bg = "#fdecea"
        icon = "❌"

    return f"""
    <div style='border: 2px solid {color}; border-radius: 10px; padding: 18px; background-color: {bg}; margin-bottom: 12px;'>
        <h4 style='color:{color}; margin:0 0 8px 0;'>{icon} 현재 달성 등급</h4>
        <p style='font-size:24px; font-weight:bold; color:{color}; margin:0;'>{level_name}</p>
    </div>
    """


def gap_box(title, amount):
    if amount > 0:
        color = "#e6f4ea"
        txt = "#0c6b2c"
        sym = f"+{amount:,.0f} 원 초과"
    elif amount < 0:
        color = "#fdecea"
        txt = "#b80000"
        sym = f"{amount:,.0f} 원 부족"
    else:
        color = "#f3f3f3"
        txt = "#000000"
        sym = "기준 달성"

    return f"""
    <div style='border: 1px solid {txt}; border-radius: 8px; background-color: {color}; padding: 12px; margin: 10px 0;'>
        <strong style='color:{txt};'>{title}: {sym}</strong>
    </div>
    """


def req_box(title, ok):
    color = "#e6f4ea" if ok else "#fdecea"
    txt = "#0c6b2c" if ok else "#b80000"
    mark_txt = "✅ 충족" if ok else "❌ 미충족"

    return f"""
    <div style='border: 1px solid {txt}; border-radius: 8px; background-color: {color}; padding: 12px; margin: 10px 0;'>
        <strong style='color:{txt};'>{title}: {mark_txt}</strong>
    </div>
    """


def make_group(df: pd.DataFrame) -> pd.DataFrame:
    rows = []

    for collector, sub in df.groupby("수금자명", dropna=False):
        req = check_convention_requirements(sub)

        rows.append({
            "수금자명": collector,
            "건수": len(sub),
            "실적보험료합계": sub["실적보험료"].sum(),
            "컨벤션환산합계": sub["컨벤션환산금액"].sum(),
            "달성등급": req["달성등급"],
            "필수조건": mark(req["필수조건"]),
            "일반": mark(req["일반달성"]),
            "더블": mark(req["더블달성"]),
            "트리플": mark(req["트리플달성"]),
            "5건": mark(req["건수5건"]),
            "한화생명5만": mark(req["한화생명5만"]),
        })

    group = pd.DataFrame(rows)

    if group.empty:
        return pd.DataFrame(columns=[
            "수금자명",
            "건수",
            "실적보험료합계",
            "컨벤션환산합계",
            "달성등급",
            "필수조건",
            "일반",
            "더블",
            "트리플",
            "5건",
            "한화생명5만",
        ])

    return group


def format_group_for_display(group: pd.DataFrame) -> pd.DataFrame:
    df = group.copy()

    for col in ["실적보험료합계", "컨벤션환산합계"]:
        if col in df.columns:
            df[col] = df[col].map(won)

    return df


# ── 엑셀 출력 ────────────────────────────────────────────────
def write_table(ws, df_for_sheet: pd.DataFrame, start_row: int = 1, name_suffix: str = "A"):
    global TABLE_SEQ

    r_idx = start_row

    for r_idx, row in enumerate(dataframe_to_rows(df_for_sheet, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    end_col_letter = ws.cell(row=start_row, column=max(df_for_sheet.shape[1], 1)).column_letter
    last_row = r_idx if df_for_sheet.shape[0] > 0 else start_row

    TABLE_SEQ += 1
    display_name = safe_table_name(f"tbl_{ws.title}_{name_suffix}_{TABLE_SEQ}")

    table = Table(displayName=display_name, ref=f"A{start_row}:{end_col_letter}{last_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(table)

    autosize_columns_full(ws, padding=5)

    return last_row


def write_title(ws, row, title):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=13)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def write_requirements_line(ws, row, dfin: pd.DataFrame):
    req = check_convention_requirements(dfin)

    line = (
        f"컨벤션 조건: "
        f"달성등급 [{req['달성등급']}]  |  "
        f"일반 {CONVENTION_GENERAL_TARGET:,.0f}원 {mark(req['일반달성'])}  |  "
        f"더블 {CONVENTION_DOUBLE_TARGET:,.0f}원 {mark(req['더블달성'])}  |  "
        f"트리플 {CONVENTION_TRIPLE_TARGET:,.0f}원 {mark(req['트리플달성'])}  |  "
        f"건수 {CONVENTION_MIN_COUNT}건 {mark(req['건수5건'])}  |  "
        f"한화생명 {CONVENTION_HANWHA_MIN_PREMIUM:,.0f}원 이상 1건 {mark(req['한화생명5만'])}  |  "
        f"필수조건 {mark(req['필수조건'])}"
    )

    cell = ws.cell(row=row, column=1, value=line)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def write_totals_block(ws, dfin: pd.DataFrame, start_row: int):
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    fill = PatternFill("solid", fgColor="F2F2F2")

    perf = dfin["실적보험료"].sum()
    conv = dfin["컨벤션환산금액"].sum()
    req = check_convention_requirements(dfin)

    rows = [
        ["실적보험료 합계", won(perf)],
        ["컨벤션 환산 합계", won(conv)],
        ["현재 달성등급", req["달성등급"]],
        ["일반 달성 목표 대비", f"{conv - CONVENTION_GENERAL_TARGET:,.0f} 원"],
        ["더블 달성 목표 대비", f"{conv - CONVENTION_DOUBLE_TARGET:,.0f} 원"],
        ["트리플 달성 목표 대비", f"{conv - CONVENTION_TRIPLE_TARGET:,.0f} 원"],
        ["계약 5건 이상", mark(req["건수5건"])],
        ["한화생명 5만원 이상 1건", mark(req["한화생명5만"])],
        ["필수조건", mark(req["필수조건"])],
    ]

    for i, row_data in enumerate(rows, start=start_row):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

            if j == 1:
                cell.fill = fill
                cell.font = Font(bold=True)

    autosize_columns_full(ws, padding=5)

    return start_row + len(rows)


def build_workbook(df: pd.DataFrame, group: pd.DataFrame, excluded_disp_all: pd.DataFrame):
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "요약"

    write_title(ws_summary, 1, "컨벤션 수금자별 요약")
    next_row = write_table(ws_summary, format_group_for_display(group), start_row=2, name_suffix="SUMMARY")

    write_requirements_line(ws_summary, next_row + 2, df)
    total_next = write_totals_block(ws_summary, df, next_row + 4)

    if not excluded_disp_all.empty:
        write_title(ws_summary, total_next + 2, "제외 계약 목록")
        write_table(ws_summary, excluded_disp_all, start_row=total_next + 3, name_suffix="EXCLUDED")

    collectors = sorted(df["수금자명"].astype(str).unique().tolist())

    for collector in collectors:
        sub = df[df["수금자명"].astype(str) == collector].copy()
        sheet_title = unique_sheet_name(wb, collector)
        ws = wb.create_sheet(title=sheet_title)

        write_title(ws, 1, f"{collector} 컨벤션 환산 결과")
        table_last_row = write_table(ws, to_styled(sub), start_row=2, name_suffix="DETAIL")

        next_row = write_totals_block(ws, sub, table_last_row + 2)
        write_requirements_line(ws, next_row + 1, sub)

        ex_sub = excluded_disp_all[
            excluded_disp_all["수금자명"].astype(str) == str(collector)
        ]

        if not ex_sub.empty:
            write_title(ws, next_row + 3, "제외 계약")
            write_table(ws, ex_sub, start_row=next_row + 4, name_suffix="EXCLUDED")

    return wb


# ── 메인 실행 ────────────────────────────────────────────────
def run():
    st.title("🏆 컨벤션 계산기")
    st.caption("컨벤션 기준으로 보험 계약 실적을 환산합니다.")

    with st.sidebar:
        st.subheader("🏆 컨벤션 기준")
        st.markdown(
            f"""
            **달성 기준**
            - 일반 달성: **{CONVENTION_GENERAL_TARGET:,.0f}원**
            - 더블 달성: **{CONVENTION_DOUBLE_TARGET:,.0f}원**
            - 트리플 달성: **{CONVENTION_TRIPLE_TARGET:,.0f}원**

            **필수 조건**
            - 계약 건수 **{CONVENTION_MIN_COUNT}건 이상**
            - 한화생명 **{CONVENTION_HANWHA_MIN_PREMIUM:,.0f}원 이상 1건**

            ※ 필수 조건 미충족 시 금액이 높아도 달성으로 인정하지 않습니다.
            """
        )

        st.markdown(
            """
            **환산율**
            - 한화생명: 150%
            - 생명보험 10년납 미만: 50%
            - 생명보험 10년납 이상: 100%
            - 손해보험: 200%
            - 흥국화재, KB손해, 한화손해, DB손해: 300%
            """
        )

    uploaded_file = st.file_uploader("📂 컨벤션 계산용 Excel 파일 업로드 (.xlsx)", type=["xlsx"])

    if not uploaded_file:
        st.info("📤 계약 목록 Excel 파일(.xlsx)을 업로드해주세요.")
        return

    base_filename = os.path.splitext(uploaded_file.name)[0]
    download_filename = f"{base_filename}_컨벤션환산결과.xlsx"

    try:
        raw = load_df(uploaded_file)
    except Exception as e:
        st.error(f"❌ 엑셀 파일을 읽는 중 오류가 발생했습니다: {e}")
        st.stop()

    df_valid, excluded_df = exclude_contracts(raw)
    excluded_disp_all = build_excluded_with_reason(excluded_df)

    missing = check_required_columns(df_valid)

    if missing:
        st.error(
            "❌ 업로드된 파일에 다음 항목이 필요합니다:\n\n"
            + ", ".join(sorted(missing))
        )
        st.stop()

    df = compute_convention(df_valid)

    invalid_dates = df[df["계약일자_raw"].isna()]

    if not invalid_dates.empty:
        st.warning(
            f"⚠️ {len(invalid_dates)}건의 계약일자가 날짜로 인식되지 않았습니다."
        )

    if not excluded_df.empty:
        st.warning(
            f"⚠️ 제외된 계약 {len(excluded_df)}건이 있습니다. "
            "제외 조건: 일시납 / 연금성·저축성 / 철회·해약·실효"
        )

        with st.expander("🚫 제외된 계약 목록 보기", expanded=False):
            st.dataframe(excluded_disp_all, use_container_width=True)

    collectors = ["전체"] + sorted(df["수금자명"].astype(str).unique().tolist())

    selected_collector = st.selectbox("👤 수금자명 선택", collectors, index=0)

    show_df = (
        df
        if selected_collector == "전체"
        else df[df["수금자명"].astype(str) == selected_collector].copy()
    )

    st.subheader(f"📄 {'전체' if selected_collector == '전체' else selected_collector} 컨벤션 환산 결과")
    st.dataframe(to_styled(show_df), use_container_width=True)

    req = check_convention_requirements(show_df)
    conv_sum = req["컨벤션환산금액"]

    st.subheader("🏆 컨벤션 달성 현황")

    st.markdown(
        money_box("컨벤션 환산보험료 합계", conv_sum),
        unsafe_allow_html=True,
    )

    st.markdown(
        level_box(req["달성등급"]),
        unsafe_allow_html=True,
    )

    if req["달성등급"] == "필수조건 미충족":
        st.warning(
            f"금액 기준으로는 [{req['금액수준']}] 수준이지만, "
            "계약 5건 이상 또는 한화생명 5만원 이상 1건 조건이 충족되지 않아 "
            "최종 달성으로 인정되지 않습니다."
        )

    st.markdown(
        gap_box(
            f"일반 달성 {CONVENTION_GENERAL_TARGET:,.0f}원 대비",
            conv_sum - CONVENTION_GENERAL_TARGET,
        ),
        unsafe_allow_html=True,
    )

    st.markdown(
        gap_box(
            f"더블 달성 {CONVENTION_DOUBLE_TARGET:,.0f}원 대비",
            conv_sum - CONVENTION_DOUBLE_TARGET,
        ),
        unsafe_allow_html=True,
    )

    st.markdown(
        gap_box(
            f"트리플 달성 {CONVENTION_TRIPLE_TARGET:,.0f}원 대비",
            conv_sum - CONVENTION_TRIPLE_TARGET,
        ),
        unsafe_allow_html=True,
    )

    st.markdown(
        req_box(f"계약 건수 {CONVENTION_MIN_COUNT}건 이상", req["건수5건"]),
        unsafe_allow_html=True,
    )

    st.markdown(
        req_box(
            f"한화생명 {CONVENTION_HANWHA_MIN_PREMIUM:,.0f}원 이상 1건",
            req["한화생명5만"],
        ),
        unsafe_allow_html=True,
    )

    st.markdown(
        req_box("컨벤션 필수 조건", req["필수조건"]),
        unsafe_allow_html=True,
    )

    st.markdown(
        req_box("일반 달성", req["일반달성"]),
        unsafe_allow_html=True,
    )

    st.markdown(
        req_box("더블 달성", req["더블달성"]),
        unsafe_allow_html=True,
    )

    st.markdown(
        req_box("트리플 달성", req["트리플달성"]),
        unsafe_allow_html=True,
    )

    st.subheader("🧮 수금자명별 요약")

    group = make_group(df)
    st.dataframe(format_group_for_display(group), use_container_width=True)

    wb = build_workbook(df, group, excluded_disp_all)

    excel_output = BytesIO()
    wb.save(excel_output)
    excel_output.seek(0)

    st.download_button(
        label="📥 컨벤션 환산 결과 엑셀 다운로드",
        data=excel_output,
        file_name=download_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )